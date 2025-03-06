from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import os
from os import listdir
from os.path import exists
from os import makedirs

# 채널명 정제 함수
def arrangeChannel(channel):
  if channel == "jja6806(옥션)":
    return "옥션"
  elif channel == "jja6806(지마켓)":
    return "지마켓"
  elif channel == "카카오스타일(저스틴23)":
    return "카카오스타일"
  elif channel == "위메프(저스틴23)":
    return "위메프"
  elif channel == "11번가(jja6806)":
    return "11번가"
  elif channel == "티몬(저스틴23)":
    return "티몬"
  elif channel == "스마트스토어(저스틴23)":
    return "스마트스토어"
  elif channel == "톡스토어(저스틴23)":
    return "톡스토어"
  elif channel == "톡스토어(저스틴23)_계정미사용":
    return "톡스토어"
  elif channel == "하프클럽":
    return "보리보리"
  elif channel == "키즈노트(저스틴23)":
    return "키즈노트"
  else:
    return channel

# 폴더 내 엑셀 파일 검색
currPath = os.getcwd()
files = listdir(currPath + '\\주문건 정리본')
excelFileList = []

for i in files:
  if(i.split('.')[-1] == 'xlsx'):
    if(not i.startswith('~')):
      excelFileList.append(i)

date = input("대상 연월일을 입력하세요(ex. 20240101) : ")

dailyWb = Workbook()
dailyWs = dailyWb.active
dailyWs.title = "판매량 체크"

fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True)

dailyWs.cell(1, 1).value = '주문건수(종합)'
dailyWs.cell(1, 2).value = '판매량'
dailyWs.cell(1, 4).value = '주문건수(상품기준)'
dailyWs.cell(1, 5).value = '판매량'
dailyWs.cell(1, 7).value = '주문건수(사이즈기준)'
dailyWs.cell(1, 8).value = '판매량'
dailyWs.cell(1, 10).value = '주문건수(채널기준)'
dailyWs.cell(1, 11).value = '판매량'
dailyWs.cell(1, 13).value = '주문건수(주소기준)'
dailyWs.cell(1, 14).value = '판매량'
dailyWs.cell(1, 16).value = '주문건수(주소고객기준)'
dailyWs.cell(1, 17).value = '판매량'
dailyWs.cell(1, 19).value = '주문건수(주문수량기준)'
dailyWs.cell(1, 20).value = '판매량'
dailyWs.cell(1, 22).value = '상품 판매채널 종합'
dailyWs.cell(1, 23).value = '판매채널'
dailyWs.cell(1, 25).value = '상품(상세) 판매채널 종합'
dailyWs.cell(1, 26).value = '판매채널'
dailyWs.cell(1, 28).value = '상품 주문번호 종합'
dailyWs.cell(1, 29).value = '주문번호'
dailyWs.cell(1, 31).value = '상품(상세) 주문번호 종합'
dailyWs.cell(1, 32).value = '주문번호'
dailyWs.cell(1, 34).value = '주문건수(사이즈기준 - 채널별 취합용)'
dailyWs.cell(1, 35).value = '판매량'
dailyWs.cell(1, 37).value = '주문건수(주문수량기준 - 채널별 취합용)'
dailyWs.cell(1, 38).value = '판매량'

dailyWs.cell(1, 1).alignment = fillAlignment
dailyWs.cell(1, 2).alignment = fillAlignment
dailyWs.cell(1, 4).alignment = fillAlignment
dailyWs.cell(1, 5).alignment = fillAlignment
dailyWs.cell(1, 7).alignment = fillAlignment
dailyWs.cell(1, 8).alignment = fillAlignment
dailyWs.cell(1, 10).alignment = fillAlignment
dailyWs.cell(1, 11).alignment = fillAlignment
dailyWs.cell(1, 13).alignment = fillAlignment
dailyWs.cell(1, 14).alignment = fillAlignment
dailyWs.cell(1, 16).alignment = fillAlignment
dailyWs.cell(1, 17).alignment = fillAlignment
dailyWs.cell(1, 19).alignment = fillAlignment
dailyWs.cell(1, 20).alignment = fillAlignment
dailyWs.cell(1, 22).alignment = fillAlignment
dailyWs.cell(1, 23).alignment = fillAlignment
dailyWs.cell(1, 25).alignment = fillAlignment
dailyWs.cell(1, 26).alignment = fillAlignment
dailyWs.cell(1, 28).alignment = fillAlignment
dailyWs.cell(1, 29).alignment = fillAlignment
dailyWs.cell(1, 31).alignment = fillAlignment
dailyWs.cell(1, 32).alignment = fillAlignment
dailyWs.cell(1, 34).alignment = fillAlignment
dailyWs.cell(1, 35).alignment = fillAlignment
dailyWs.cell(1, 37).alignment = fillAlignment
dailyWs.cell(1, 38).alignment = fillAlignment

dailyWs.cell(1, 1).font = fillFont
dailyWs.cell(1, 2).font = fillFont
dailyWs.cell(1, 4).font = fillFont
dailyWs.cell(1, 5).font = fillFont
dailyWs.cell(1, 7).font = fillFont
dailyWs.cell(1, 8).font = fillFont
dailyWs.cell(1, 10).font = fillFont
dailyWs.cell(1, 11).font = fillFont
dailyWs.cell(1, 13).font = fillFont
dailyWs.cell(1, 14).font = fillFont
dailyWs.cell(1, 16).font = fillFont
dailyWs.cell(1, 17).font = fillFont
dailyWs.cell(1, 19).font = fillFont
dailyWs.cell(1, 20).font = fillFont
dailyWs.cell(1, 22).font = fillFont
dailyWs.cell(1, 23).font = fillFont
dailyWs.cell(1, 25).font = fillFont
dailyWs.cell(1, 26).font = fillFont
dailyWs.cell(1, 28).font = fillFont
dailyWs.cell(1, 29).font = fillFont
dailyWs.cell(1, 31).font = fillFont
dailyWs.cell(1, 32).font = fillFont
dailyWs.cell(1, 34).font = fillFont
dailyWs.cell(1, 35).font = fillFont
dailyWs.cell(1, 37).font = fillFont
dailyWs.cell(1, 38).font = fillFont

dailyWs.cell(1, 1).fill = fillData
dailyWs.column_dimensions['A'].width = 40
dailyWs.column_dimensions['B'].width = 10
dailyWs.column_dimensions['C'].width = 10
dailyWs.column_dimensions['D'].width = 30
dailyWs.column_dimensions['E'].width = 10
dailyWs.column_dimensions['F'].width = 10
dailyWs.column_dimensions['G'].width = 20
dailyWs.column_dimensions['H'].width = 10
dailyWs.column_dimensions['J'].width = 20
dailyWs.column_dimensions['K'].width = 10
dailyWs.column_dimensions['M'].width = 30
dailyWs.column_dimensions['N'].width = 10
dailyWs.column_dimensions['P'].width = 30
dailyWs.column_dimensions['Q'].width = 10
dailyWs.column_dimensions['S'].width = 30
dailyWs.column_dimensions['T'].width = 10
dailyWs.column_dimensions['V'].width = 30
dailyWs.column_dimensions['W'].width = 50
dailyWs.column_dimensions['Y'].width = 30
dailyWs.column_dimensions['Z'].width = 50
dailyWs.column_dimensions['AB'].width = 30
dailyWs.column_dimensions['AC'].width = 50
dailyWs.column_dimensions['AE'].width = 30
dailyWs.column_dimensions['AF'].width = 50
dailyWs.column_dimensions['AH'].width = 40
dailyWs.column_dimensions['AI'].width = 10
dailyWs.column_dimensions['AK'].width = 40
dailyWs.column_dimensions['AL'].width = 10

fillData2 = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
dailyWs["A1"].fill = fillData2
dailyWs["B1"].fill = fillData2
dailyWs["D1"].fill = fillData2
dailyWs["E1"].fill = fillData2
dailyWs["G1"].fill = fillData2
dailyWs["H1"].fill = fillData2
dailyWs["J1"].fill = fillData2
dailyWs["K1"].fill = fillData2
dailyWs["M1"].fill = fillData2
dailyWs["N1"].fill = fillData2
dailyWs["P1"].fill = fillData2
dailyWs["Q1"].fill = fillData2
dailyWs["S1"].fill = fillData2
dailyWs["T1"].fill = fillData2
dailyWs["V1"].fill = fillData2
dailyWs["W1"].fill = fillData2
dailyWs["Y1"].fill = fillData2
dailyWs["Z1"].fill = fillData2
dailyWs["AB1"].fill = fillData2
dailyWs["AC1"].fill = fillData2
dailyWs["AE1"].fill = fillData2
dailyWs["AF1"].fill = fillData2
dailyWs["AH1"].fill = fillData2
dailyWs["AI1"].fill = fillData2
dailyWs["AK1"].fill = fillData2
dailyWs["AL1"].fill = fillData2

orderDict = {}
orderDictPrd = {}
orderDictSize = {}
orderDictSizeChannelAcc = {}
orderDictChannel = {}
orderDictAddress = {}
orderDictCustomer = {}
orderDictQuantity = {}
orderDictQuantityChannelAcc = {}
orderDictPrdNums = {}
orderDictPrdDetailNums = {}

# 제품별 판매채널 정보
orderDictPrdChannel = {}

# 제품별(상세) 판매채널 정보
orderDictPrdDetailChannel = {}

for file in excelFileList:

  wb = load_workbook(currPath + '\\주문건 정리본\\' + file)
  ws = wb['추출내용']

  first_row = 2
  last_row = ws.max_row + 1
  dailyWs_last_row = dailyWs.max_row + 1

  for i in range(first_row, last_row):
    if ws.cell(row=i, column=9).value == None or ws.cell(row=i, column=9).value == '':
      continue
    else:
      print(ws.cell(row=i, column=9).value + " / 판매수량 : " + str(ws.cell(row=i, column=10).value))

  for i in range(first_row, last_row):
    if ws.cell(i, 9).value == None or ws.cell(i, 9).value == '':
      continue
    else:
      if ws.cell(i, 9).value not in orderDict:
        orderDict[ws.cell(i, 9).value] = ws.cell(i, 10).value
      else:
        orderDict[ws.cell(i, 9).value] += ws.cell(i, 10).value
        
  for i in range(first_row, last_row):
    if ws.cell(i, 12).value == None or ws.cell(i, 12).value == '':
      continue
    else:    
      if ws.cell(i, 12).value not in orderDictPrd:
        orderDictPrd[ws.cell(i, 12).value] = ws.cell(i, 13).value
      else:
        orderDictPrd[ws.cell(i, 12).value] += ws.cell(i, 13).value
  
  for i in range(first_row, last_row):
    if ws.cell(i, 15).value == None or ws.cell(i, 15).value == '':
      continue
    else:
      if ws.cell(i, 15).value not in orderDictSize:
        orderDictSize[ws.cell(i, 15).value] = ws.cell(i, 16).value
      else:
        orderDictSize[ws.cell(i, 15).value] += ws.cell(i, 16).value
  
  for i in range(first_row, last_row):
    if ws.cell(i, 36).value == None or ws.cell(i, 36).value == '':
      continue
    else:
      if ws.cell(i, 36).value not in orderDictSizeChannelAcc:
        orderDictSizeChannelAcc[ws.cell(i, 36).value] = ws.cell(i, 37).value
      else:
        orderDictSizeChannelAcc[ws.cell(i, 36).value] += ws.cell(i, 37).value
        
  for i in range(first_row, last_row):
    if ws.cell(i, 18).value == None or ws.cell(i, 18).value == '':
      continue
    else:
      if ws.cell(i, 18).value not in orderDictChannel:
        orderDictChannel[ws.cell(i, 18).value] = ws.cell(i, 19).value
      else:
        orderDictChannel[ws.cell(i, 18).value] += ws.cell(i, 19).value
        
  for i in range(first_row, last_row):
    if ws.cell(i, 21).value == None or ws.cell(i, 21).value == '':
      continue
    else:
      if ws.cell(i, 21).value not in orderDictAddress:
        orderDictAddress[ws.cell(i, 21).value] = ws.cell(i, 22).value
      else:
        orderDictAddress[ws.cell(i, 21).value] += ws.cell(i, 22).value
        
  for i in range(first_row, last_row):
    if ws.cell(i, 24).value == None or ws.cell(i, 24).value == '':
      continue
    else:
      if ws.cell(i, 24).value not in orderDictCustomer:
        orderDictCustomer[ws.cell(i, 24).value] = ws.cell(i, 25).value
      else:
        orderDictCustomer[ws.cell(i, 24).value] += ws.cell(i, 25).value
        
  for i in range(first_row, last_row):
    if ws.cell(i, 27).value == None or ws.cell(i, 27).value == '':
      continue
    else:
      if ws.cell(i, 27).value not in orderDictQuantity:
        orderDictQuantity[ws.cell(i, 27).value] = ws.cell(i, 28).value
      else:
        orderDictQuantity[ws.cell(i, 27).value] += ws.cell(i, 28).value
        
  for i in range(first_row, last_row):
    if ws.cell(i, 39).value == None or ws.cell(i, 39).value == '':
      continue
    else:
      if ws.cell(i, 39).value not in orderDictQuantityChannelAcc:
        orderDictQuantityChannelAcc[ws.cell(i, 39).value] = ws.cell(i, 40).value
      else:
        orderDictQuantityChannelAcc[ws.cell(i, 39).value] += ws.cell(i, 40).value
        
  for i in range(first_row, last_row):
    if ws.cell(i, 30).value == None or ws.cell(i, 30).value == '':
      continue
    else:
      if ws.cell(i, 30).value not in orderDictPrdNums:
        orderDictPrdNums[ws.cell(i, 30).value] = ws.cell(i, 31).value
      else:
        orderDictPrdNums[ws.cell(i, 30).value] += ", " + ws.cell(i, 31).value
        
  for i in range(first_row, last_row):
    if ws.cell(i, 33).value == None or ws.cell(i, 33).value == '':
      continue
    else:
      if ws.cell(i, 33).value not in orderDictPrdDetailNums:
        orderDictPrdDetailNums[ws.cell(i, 33).value] = ws.cell(i, 34).value
      else:
        orderDictPrdDetailNums[ws.cell(i, 33).value] += ", " + ws.cell(i, 34).value
        
  for i in range(first_row, last_row):
    if ws.cell(i, 2).value == None or ws.cell(i, 2).value == '':
      continue
    else:
      if ws.cell(i, 2).value not in orderDictPrdChannel:
        orderDictPrdChannel[ws.cell(i, 2).value] = [ws.cell(i, 6).value]
      else:
        if ws.cell(i, 6).value not in orderDictPrdChannel[ws.cell(i, 2).value]:
          orderDictPrdChannel[ws.cell(i, 2).value].append(ws.cell(i, 6).value)
        
  for i in range(first_row, last_row):
    if ws.cell(i, 5).value == None or ws.cell(i, 5).value == '':
      continue
    else:
      if ws.cell(i, 5).value not in orderDictPrdDetailChannel:
        orderDictPrdDetailChannel[ws.cell(i, 5).value] = [ws.cell(i, 6).value]
      else:
        if ws.cell(i, 6).value not in orderDictPrdDetailChannel[ws.cell(i, 5).value]:
          orderDictPrdDetailChannel[ws.cell(i, 5).value].append(ws.cell(i, 6).value)
          
  orderDictCnt = 2
  for key, value in orderDict.items():
    dailyWs.cell(orderDictCnt, 1).value = key
    dailyWs.cell(orderDictCnt, 2).value = value
    orderDictCnt += 1
  
  orderDictPrdCnt = 2
  for key, value in orderDictPrd.items():
    dailyWs.cell(orderDictPrdCnt, 4).value = key
    dailyWs.cell(orderDictPrdCnt, 5).value = value
    orderDictPrdCnt += 1
  
  orderDictSizeCnt = 2
  for key, value in orderDictSize.items():
    if dailyWs.cell(orderDictSizeCnt, 7).value == '5호':
      dailyWs.cell(orderDictSizeCnt, 7).value = '05호'
    elif dailyWs.cell(orderDictSizeCnt, 7).value == '7호':
      dailyWs.cell(orderDictSizeCnt, 7).value = '07호'
    elif dailyWs.cell(orderDictSizeCnt, 7).value == '9호':
      dailyWs.cell(orderDictSizeCnt, 7).value = '09호'
    else:
      dailyWs.cell(orderDictSizeCnt, 7).value = key
    
    dailyWs.cell(orderDictSizeCnt, 8).value = value
    orderDictSizeCnt += 1
  
  orderDictChannelCnt = 2
  for key, value in orderDictChannel.items():
    dailyWs.cell(orderDictChannelCnt, 10).value = key
    dailyWs.cell(orderDictChannelCnt, 11).value = value
    orderDictChannelCnt += 1
    
  orderDictAddressCnt = 2
  for key, value in orderDictAddress.items():
    dailyWs.cell(orderDictAddressCnt, 13).value = key
    dailyWs.cell(orderDictAddressCnt, 14).value = value
    orderDictAddressCnt += 1
    
  orderDictCustomerCnt = 2
  for key, value in orderDictCustomer.items():
    dailyWs.cell(orderDictCustomerCnt, 16).value = key
    dailyWs.cell(orderDictCustomerCnt, 17).value = value
    orderDictCustomerCnt += 1
    
  orderDictQuantityCnt = 2
  for key, value in orderDictQuantity.items():
    dailyWs.cell(orderDictQuantityCnt, 19).value = key
    dailyWs.cell(orderDictQuantityCnt, 20).value = value
    orderDictQuantityCnt += 1
    
  orderDictPrdChannelCnt = 2
  for key, value in orderDictPrdChannel.items():
    valueText = ""
    dailyWs.cell(orderDictPrdChannelCnt, 22).value = key
    for idx, i in enumerate(value):
      try:
        if idx == 0:
          valueText += arrangeChannel(i)
        else:
          valueText += "/{}".format(arrangeChannel(i))
      except:
        pass
    dailyWs.cell(orderDictPrdChannelCnt, 23).value = valueText
    orderDictPrdChannelCnt += 1
    
  orderDictPrdDetailChannelCnt = 2
  for key, value in orderDictPrdDetailChannel.items():
    valueText = ""
    dailyWs.cell(orderDictPrdDetailChannelCnt, 25).value = key
    for idx, i in enumerate(value):
      try:
        if idx == 0:
          valueText += arrangeChannel(i)
        else:
          valueText += "/{}".format(arrangeChannel(i))
      except:
        pass
    dailyWs.cell(orderDictPrdDetailChannelCnt, 26).value = valueText
    orderDictPrdDetailChannelCnt += 1
  
  orderDictPrdNumsCnt = 2
  for key, value in orderDictPrdNums.items():
    dailyWs.cell(orderDictPrdNumsCnt, 28).value = key
    dailyWs.cell(orderDictPrdNumsCnt, 29).value = value
    dailyWs.cell(orderDictPrdNumsCnt, 30).value = " "
    orderDictPrdNumsCnt += 1
    
  orderDictPrdDetailNumsCnt = 2
  for key, value in orderDictPrdDetailNums.items():
    dailyWs.cell(orderDictPrdDetailNumsCnt, 31).value = key
    dailyWs.cell(orderDictPrdDetailNumsCnt, 32).value = value
    dailyWs.cell(orderDictPrdDetailNumsCnt, 33).value = " "
    orderDictPrdDetailNumsCnt += 1
    
  orderDictSizeChannelAccCnt= 2
  for key, value in orderDictSizeChannelAcc.items():
    dailyWs.cell(orderDictSizeChannelAccCnt, 34).value = key
    dailyWs.cell(orderDictSizeChannelAccCnt, 35).value = value
    dailyWs.cell(orderDictSizeChannelAccCnt, 36).value = " "
    orderDictSizeChannelAccCnt += 1
    
  orderDictQuantityChannelAccCnt= 2
  for key, value in orderDictQuantityChannelAcc.items():
    dailyWs.cell(orderDictQuantityChannelAccCnt, 37).value = key
    dailyWs.cell(orderDictQuantityChannelAccCnt, 38).value = value
    dailyWs.cell(orderDictQuantityChannelAccCnt, 39).value = " "
    orderDictQuantityChannelAccCnt += 1
    
dailyWb.save(currPath + '\\' + date + '.xlsx')

import win32com.client

excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
wb = excel.Workbooks.Open(currPath + '\\' + date + '.xlsx')
ws = wb.Worksheets('판매량 체크')

ws.Range('A:B').Sort(Key1=ws.Range('B1'), Order1=2)
ws.Range('D:E').Sort(Key1=ws.Range('E1'), Order1=2)

# 사이즈기준은 판매량에 따른 정렬 미적용
ws.Range('G:H').Sort(Key1=ws.Range('H1'), Order1=2)
ws.Range('AH:AI').Sort(Key1=ws.Range('AI1'), Order1=2)

ws.Range('J:K').Sort(Key1=ws.Range('K1'), Order1=2)
ws.Range('M:N').Sort(Key1=ws.Range('N1'), Order1=2)
ws.Range('P:Q').Sort(Key1=ws.Range('Q1'), Order1=2)

# 주문수량기준은 판매량에 따른 정렬 미적용
# ws.Range('S:T').Sort(Key1=ws.Range('T1'), Order1=2)
wb.Save()

excel.Quit()