from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import os
from os import listdir
from os.path import exists
from os import makedirs
import datetime
import json

# 폴더 내 엑셀 파일 검색
currPath = os.getcwd()
files = listdir(currPath)
excelFileList = []

for i in files:
  if(i.split('.')[-1] == 'xlsx'):
    if(not i.startswith('~')):
      excelFileList.append(i)
      
productsSeason = {} # 제품 시즌정보
productsSeasonOrder = {'F': 0, 'S': 0, 'W': 0} # 시즌별 판매정보
productsCode = {} # 제품 코드정보
productsUrl = {} # 제품 url 정보
prdF = [] # 봄가을 판매상품 정보
prdS = [] # 여름 판매상품 정보
prdW = [] # 겨울 판매상품 정보


# DB 불러오기
import sqlite3

# 상품 정보 생성
con = sqlite3.connect('D:/1.업무/10.기타자료/Development/db/productsData.db')
cur = con.cursor()
cur.execute("SELECT PrdName, prdCode, Season, Url from ProductsData")
data = cur.fetchall()
for i in data:
  productsCode[i[0]] = i[1]
  productsSeason[i[0]] = i[2]
  productsUrl[i[0]] = i[3]
  
for file in excelFileList:

  wb = load_workbook(currPath + '\\' + file)
  wb.create_sheet('상품사진삽입')

  ws_name = wb.get_sheet_names()

  sheet1 = wb[str(ws_name[0])]
  sheet2 = wb[str(ws_name[1])]
  
  wb.active = wb['상품사진삽입']
  
  fillAlignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
  fillFont = Font(bold=True)
  
  sheet2.cell(1, 1).value = '주문건수(상품기준)'
  sheet2.cell(1, 2).value = '판매량'
  sheet2.cell(1, 3).value = '대표이미지'
  sheet2.cell(1, 5).value = '봄가을'
  sheet2.cell(1, 6).value = '여름'
  sheet2.cell(1, 7).value = '겨울'
  sheet2.cell(1, 10).value = '상품명'
  sheet2.cell(1, 11).value = '사이즈'
  sheet2.cell(1, 12).value = '판매량'
  sheet2.cell(1, 1).alignment = fillAlignment
  sheet2.cell(1, 2).alignment = fillAlignment
  sheet2.cell(1, 3).alignment = fillAlignment
  sheet2.cell(1, 5).alignment = fillAlignment
  sheet2.cell(1, 6).alignment = fillAlignment
  sheet2.cell(1, 7).alignment = fillAlignment
  sheet2.cell(1, 10).alignment = fillAlignment
  sheet2.cell(1, 11).alignment = fillAlignment
  sheet2.cell(1, 12).alignment = fillAlignment
  sheet2.cell(2, 5).alignment = fillAlignment
  sheet2.cell(2, 6).alignment = fillAlignment
  sheet2.cell(2, 7).alignment = fillAlignment
  sheet2.cell(2, 8).alignment = fillAlignment
  sheet2.cell(3, 5).alignment = fillAlignment
  sheet2.cell(3, 6).alignment = fillAlignment
  sheet2.cell(3, 7).alignment = fillAlignment
  sheet2.cell(1, 1).font = fillFont
  sheet2.cell(1, 2).font = fillFont
  sheet2.cell(1, 3).font = fillFont
  sheet2.cell(1, 5).font = fillFont
  sheet2.cell(1, 6).font = fillFont
  sheet2.cell(1, 7).font = fillFont
  sheet2.cell(1, 10).font = fillFont
  sheet2.cell(1, 11).font = fillFont
  sheet2.cell(1, 12).font = fillFont
  sheet2.cell(2, 8).font = fillFont

  sheet2.column_dimensions['A'].width = 40
  sheet2.column_dimensions['B'].width = 10
  sheet2.column_dimensions['C'].width = 12
  sheet2.column_dimensions['E'].width = 12
  sheet2.column_dimensions['F'].width = 12
  sheet2.column_dimensions['G'].width = 12
  sheet2.column_dimensions['H'].width = 12
  sheet2.column_dimensions['J'].width = 25

  fillData = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
  fillData2 = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
  
  sheet2["A1"].fill = fillData
  sheet2["B1"].fill = fillData
  sheet2["C1"].fill = fillData
  sheet2["E1"].fill = fillData
  sheet2["F1"].fill = fillData
  sheet2["G1"].fill = fillData
  sheet2["J1"].fill = fillData
  sheet2["K1"].fill = fillData
  sheet2["L1"].fill = fillData

  for sheet in wb:
    if sheet.title == '상품사진삽입':
      sheet.sheet_view.tabSelected = True
    else:
      sheet.sheet_view.tabSelected = False
      
  first_row = 2
  last_row = sheet1.max_row + 1  
  
  for i in range(first_row, last_row):
    try:
      sheet1.cell(row=i, column=24).value = " "
      sheet1.cell(row=i, column=27).value = " "
      sheet1.cell(row=i, column=30).value = " "
      sheet1.cell(row=i, column=33).value = " "
      
      if sheet1.cell(row=i, column=4).value == None or sheet1.cell(row=i, column=4).value == '':
        continue
      else:
        sheet2["A" + str(i)].alignment = Alignment(vertical='center')
        sheet2["B" + str(i)].alignment = Alignment(vertical='center')
        # print(sheet1.cell(row=i, column=4).value + " / 판매수량 : " + str(sheet1.cell(row=i, column=5).value))
        sheet2.cell(row=i, column=1).value = sheet1.cell(row=i, column=4).value
        sheet2.cell(row=i, column=2).value = sheet1.cell(row=i, column=5).value
        
        productsSeasonOrder[productsSeason[sheet1.cell(row=i, column=4).value]] += sheet1.cell(row=i, column=5).value
        if productsSeason[sheet1.cell(row=i, column=4).value] == 'F':
          prdF.append(sheet1.cell(row=i, column=4).value)
        elif productsSeason[sheet1.cell(row=i, column=4).value] == 'S':
          prdS.append(sheet1.cell(row=i, column=4).value)
        else:
          prdW.append(sheet1.cell(row=i, column=4).value)
          
        try:
          sheet2.row_dimensions[i].height = 75
          # image_path = 'https://gi.esmplus.com/jja6806/thumbnail/{}.jpg'.format(sheet2.cell(row=i, column=3).value)
          # image_path = '.\\data\\images\\' + str(productsCode[sheet2.cell(row=i, column=1).value]) + '.jpg'
          image_path = '\\\\Main-pc\\KIDS_CS\\대표이미지\\resize\\' + str(sheet2.cell(row=i, column=1).value) + '.jpg'
          image = Image(image_path)
          image.width = 100
          image.height = 100
          sheet2.add_image(image, anchor='C'+str(i))
        except Exception as e:
          print(e)
          sheet2.cell(i, 1).fill = fillData2
          sheet2.cell(i, 2).fill = fillData2
          sheet2.cell(i, 3).fill = fillData2
          
    except Exception as e:
      print(e)
      
  for i in range(first_row, last_row):
    # 상품 호수별 판매량 정리 
    if sheet1.cell(row=i, column=1).value == None or sheet1.cell(row=i, column=1).value == '':
      continue
    else:
      sheet2.cell(row=i, column=10).value = sheet1.cell(row=i, column=1).value.split('/')[0]
      sheet2.cell(row=i, column=11).value = sheet1.cell(row=i, column=1).value.split('/')[-1]
      sheet2.cell(row=i, column=12).value = sheet1.cell(row=i, column=2).value
      
      sheet2.cell(row=i, column=10).alignment = Alignment(vertical='center', wrap_text=True)
      sheet2.cell(row=i, column=11).alignment = Alignment(vertical='center', wrap_text=True)
      sheet2.cell(row=i, column=12).alignment = Alignment(vertical='center', wrap_text=True)
  
  seasonOrderSum = 0
  for key, value in productsSeasonOrder.items():
    seasonOrderSum += value
  
  sheet2.cell(2, 5).value = '{0}\n({1:.1f}%)'.format(productsSeasonOrder['F'], round((productsSeasonOrder['F'] / seasonOrderSum), 3) * 100)
  sheet2.cell(2, 6).value = '{0}\n({1:.1f}%)'.format(productsSeasonOrder['S'], round((productsSeasonOrder['S'] / seasonOrderSum), 3) * 100)
  sheet2.cell(2, 7).value = '{0}\n({1:.1f}%)'.format(productsSeasonOrder['W'], round((productsSeasonOrder['W'] / seasonOrderSum), 3) * 100)
  sheet2.cell(2, 8).value = '{0}'.format(seasonOrderSum)
  
  sheet2.cell(3, 5).value = '/'.join(prdF)
  sheet2.cell(3, 6).value = '/'.join(prdS)
  sheet2.cell(3, 7).value = '/'.join(prdW)
  sheet2.cell(3, 4).value = " "
  sheet2.cell(3, 8).value = " "
  
  wb.save(currPath + '\\' + file)
  
  # 채널명 정제 함수
  def arrangeChannel(channel):
    if channel == "jja6806(옥션)":
      return "옥션"
    elif channel == "jja6806(지마켓)":
      return "지마켓"
    elif channel == "카카오스타일(저스틴23)":
      return "카카오스타일"
    elif channel == "위메프(저스틴23)":
      return "위메프"
    elif channel == "11번가(jja6806)":
      return "11번가"
    elif channel == "티몬(저스틴23)":
      return "티몬"
    elif channel == "스마트스토어(저스틴23)":
      return "스마트스토어"
    elif channel == "톡스토어(저스틴23)":
      return "톡스토어"
    elif channel == "하프클럽":
      return "보리보리"
    elif channel == "키즈노트(저스틴23)":
      return "키즈노트"
    elif channel == "네오스토어_이몰":
      return "이몰"
    else:
      return channel
  
  jsonData = {}
  jsonData['data'] = []
  jsonUrlData = {}
  
  salesInfoSize = {}
  salesInfoSizeChannelAcc = {}
  salesInfoChannel = {}
  salesInfoQnt = {}
  salesInfoQntChannelAcc = {}
  
  for i in range(first_row, last_row):
    if sheet1.cell(row=i, column=7).value == None or sheet1.cell(row=i, column=7).value == '':
        continue
    else:
      salesInfoSize[sheet1.cell(row=i, column=7).value.replace('05호', '5호').replace('07호', '7호').replace('09호', '9호')] = sheet1.cell(row=i, column=8).value
  
  for i in range(first_row, last_row):
    if sheet1.cell(row=i, column=34).value == None or sheet1.cell(row=i, column=34).value == '':
        continue
    else:
      salesInfoSizeChannelAcc[sheet1.cell(row=i, column=34).value.replace('05호', '5호').replace('07호', '7호').replace('09호', '9호')] = sheet1.cell(row=i, column=35).value
  
  for i in range(first_row, last_row):
    if sheet1.cell(row=i, column=10).value == None or sheet1.cell(row=i, column=10).value == '':
        continue
    else:
      salesInfoChannel[arrangeChannel(sheet1.cell(row=i, column=10).value)] = sheet1.cell(row=i, column=11).value
  
  for i in range(first_row, last_row):
    if sheet1.cell(row=i, column=19).value == None or sheet1.cell(row=i, column=19).value == '':
        continue
    else:
      salesInfoQnt[sheet1.cell(row=i, column=19).value] = sheet1.cell(row=i, column=20).value
  
  for i in range(first_row, last_row):
    if sheet1.cell(row=i, column=37).value == None or sheet1.cell(row=i, column=37).value == '':
        continue
    else:
      salesInfoQntChannelAcc[sheet1.cell(row=i, column=37).value] = sheet1.cell(row=i, column=38).value
      
  try:
    jsonData['data'].append({
      'checkDate': datetime.datetime.now().strftime('%Y-%m-%d'),
      'salesTotal': seasonOrderSum,
      'salesF': productsSeasonOrder['F'],
      'salesFRatio': '{0:.1f}'.format(round((productsSeasonOrder['F'] / seasonOrderSum), 3) * 100),
      'salesS': productsSeasonOrder['S'],
      'salesSRatio': '{0:.1f}'.format(round((productsSeasonOrder['S'] / seasonOrderSum), 3) * 100),
      'salesW': productsSeasonOrder['W'],
      'salesWRatio': '{0:.1f}'.format(round((productsSeasonOrder['W'] / seasonOrderSum), 3) * 100),
      'salesInfoSize': salesInfoSize,
      'salesInfoChannel': salesInfoChannel,
      'salesInfoSizeChannelAcc': salesInfoSizeChannelAcc,
      'salesInfoQnt': salesInfoQnt,
      'salesInfoQntChannelAcc': salesInfoQntChannelAcc
    })
  except Exception as e:
    print(e)

  # 일일 판매데이터 json 생성
  with open('./saleDaily_{}.json'.format(file.replace('.xlsx', '')), 'w', encoding='UTF-8') as outfile:
    json.dump(jsonData, outfile, indent=2, ensure_ascii=False)
    
  # 상품별 상세이미지 url 정보 json 생성
  with open('./prdUrlInfo.json', 'w', encoding='UTF-8') as outfile:
    json.dump(productsUrl, outfile, indent=2, ensure_ascii=False)