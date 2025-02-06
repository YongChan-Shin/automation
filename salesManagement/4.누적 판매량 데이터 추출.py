from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import os
from os import listdir
from os.path import exists
from os import makedirs

# 폴더 내 엑셀 파일 검색
currPath = os.getcwd()
files = listdir(currPath + '\\누적 판매량 산출')
excelFileList = []

for i in files:
  if(i.split('.')[-1] == 'xlsx'):
    if(not i.startswith('~')):
      excelFileList.append(i)
fileName = "{}~{}_누적판매량".format(excelFileList[0].split('.')[0], excelFileList[-1].split('.')[0])

dailyWb = Workbook()
accWs = dailyWb.active
accWs.title = "판매량 체크"

fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True)

accWs.cell(1, 1).value = '주문건수(종합)'
accWs.cell(1, 2).value = '판매량'
accWs.cell(1, 4).value = '주문건수(상품기준)'
accWs.cell(1, 5).value = '판매량'
accWs.cell(1, 7).value = '주문건수(사이즈기준)'
accWs.cell(1, 8).value = '판매량'
accWs.cell(1, 10).value = '주문건수(채널기준)'
accWs.cell(1, 11).value = '판매량'
accWs.cell(1, 13).value = '주문건수(주소기준)'
accWs.cell(1, 14).value = '판매량'
accWs.cell(1, 16).value = '주문건수(주문고객기준)'
accWs.cell(1, 17).value = '판매량'
accWs.cell(1, 19).value = '주문건수(주문수량기준)'
accWs.cell(1, 20).value = '판매량'
accWs.cell(1, 22).value = '상품 판매채널 종합'
accWs.cell(1, 23).value = '판매채널'
accWs.cell(1, 25).value = '상품(상세) 판매채널 종합'
accWs.cell(1, 26).value = '판매채널'
accWs.cell(1, 28).value = '상품 주문번호 종합'
accWs.cell(1, 29).value = '주문번호'
accWs.cell(1, 31).value = '상품(상세) 주문번호 종합'
accWs.cell(1, 32).value = '주문번호'

accWs.cell(1, 1).alignment = fillAlignment
accWs.cell(1, 2).alignment = fillAlignment
accWs.cell(1, 4).alignment = fillAlignment
accWs.cell(1, 5).alignment = fillAlignment
accWs.cell(1, 7).alignment = fillAlignment
accWs.cell(1, 8).alignment = fillAlignment
accWs.cell(1, 10).alignment = fillAlignment
accWs.cell(1, 11).alignment = fillAlignment
accWs.cell(1, 13).alignment = fillAlignment
accWs.cell(1, 14).alignment = fillAlignment
accWs.cell(1, 16).alignment = fillAlignment
accWs.cell(1, 17).alignment = fillAlignment
accWs.cell(1, 19).alignment = fillAlignment
accWs.cell(1, 20).alignment = fillAlignment
accWs.cell(1, 22).alignment = fillAlignment
accWs.cell(1, 23).alignment = fillAlignment
accWs.cell(1, 25).alignment = fillAlignment
accWs.cell(1, 26).alignment = fillAlignment
accWs.cell(1, 28).alignment = fillAlignment
accWs.cell(1, 29).alignment = fillAlignment
accWs.cell(1, 31).alignment = fillAlignment
accWs.cell(1, 32).alignment = fillAlignment

accWs.cell(1, 1).font = fillFont
accWs.cell(1, 2).font = fillFont
accWs.cell(1, 4).font = fillFont
accWs.cell(1, 5).font = fillFont
accWs.cell(1, 7).font = fillFont
accWs.cell(1, 8).font = fillFont
accWs.cell(1, 10).font = fillFont
accWs.cell(1, 11).font = fillFont
accWs.cell(1, 13).font = fillFont
accWs.cell(1, 14).font = fillFont
accWs.cell(1, 16).font = fillFont
accWs.cell(1, 17).font = fillFont
accWs.cell(1, 19).font = fillFont
accWs.cell(1, 20).font = fillFont
accWs.cell(1, 22).font = fillFont
accWs.cell(1, 23).font = fillFont
accWs.cell(1, 25).font = fillFont
accWs.cell(1, 26).font = fillFont
accWs.cell(1, 28).font = fillFont
accWs.cell(1, 29).font = fillFont
accWs.cell(1, 31).font = fillFont
accWs.cell(1, 32).font = fillFont

accWs.column_dimensions['A'].width = 40
accWs.column_dimensions['B'].width = 10
accWs.column_dimensions['C'].width = 10
accWs.column_dimensions['D'].width = 30
accWs.column_dimensions['E'].width = 10
accWs.column_dimensions['F'].width = 10
accWs.column_dimensions['G'].width = 20
accWs.column_dimensions['H'].width = 10
accWs.column_dimensions['J'].width = 20
accWs.column_dimensions['K'].width = 10
accWs.column_dimensions['M'].width = 30
accWs.column_dimensions['N'].width = 10
accWs.column_dimensions['P'].width = 30
accWs.column_dimensions['Q'].width = 10
accWs.column_dimensions['S'].width = 30
accWs.column_dimensions['T'].width = 10
accWs.column_dimensions['V'].width = 30
accWs.column_dimensions['W'].width = 50
accWs.column_dimensions['Y'].width = 30
accWs.column_dimensions['Z'].width = 50
accWs.column_dimensions['AB'].width = 30
accWs.column_dimensions['AC'].width = 50
accWs.column_dimensions['AE'].width = 30
accWs.column_dimensions['AF'].width = 50

fillData2 = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
accWs["A1"].fill = fillData2
accWs["B1"].fill = fillData2
accWs["D1"].fill = fillData2
accWs["E1"].fill = fillData2
accWs["G1"].fill = fillData2
accWs["H1"].fill = fillData2
accWs["J1"].fill = fillData2
accWs["K1"].fill = fillData2
accWs["M1"].fill = fillData2
accWs["N1"].fill = fillData2
accWs["P1"].fill = fillData2
accWs["Q1"].fill = fillData2
accWs["S1"].fill = fillData2
accWs["T1"].fill = fillData2
accWs["V1"].fill = fillData2
accWs["W1"].fill = fillData2
accWs["Y1"].fill = fillData2
accWs["Z1"].fill = fillData2
accWs["AB1"].fill = fillData2
accWs["AC1"].fill = fillData2
accWs["AE1"].fill = fillData2
accWs["AF1"].fill = fillData2

orderDict = {}
orderDictPrd = {}
orderDictSize = {}
orderDictChannel = {}
orderDictAddress = {}
orderDictCustomer = {}
orderDictQuantity = {
  '1개': 0,
  '2개': 0,
  '3개': 0,
  '4개': 0,
  '5개 이상': 0,
  '10개 이상': 0,
  '15개 이상': 0,
  '20개 이상': 0,
  '30개 이상': 0,
  '50개 이상': 0,
  '100개 이상': 0,
}
orderDictPrdNums = {}
orderDictPrdDetailNums = {}
orderDictPrdChannel = {}
orderDictPrdDetailChannel = {}

for file in excelFileList:

  wb = load_workbook(currPath + '\\누적 판매량 산출\\' + file)
  ws = wb['판매량 체크']

  first_row = 2
  last_row = ws.max_row + 1
  accWs_last_row = accWs.max_row + 1

  for i in range(first_row, last_row):
    if ws.cell(row=i, column=1).value == None or ws.cell(row=i, column=1).value == '':
      continue
    else:
      pass
      # print(ws.cell(row=i, column=1).value + " / 판매수량 : " + str(ws.cell(row=i, column=2).value))

  for i in range(first_row, last_row):
    if ws.cell(i, 1).value == None or ws.cell(i, 1).value == '':
      continue
    else:
      if ws.cell(i, 1).value not in orderDict:
        orderDict[ws.cell(i, 1).value] = ws.cell(i, 2).value
      else:
        orderDict[ws.cell(i, 1).value] += ws.cell(i, 2).value
        
  for i in range(first_row, last_row):
    if ws.cell(i, 4).value == None or ws.cell(i, 4).value == '':
      continue
    else:    
      if ws.cell(i, 4).value not in orderDictPrd:
        orderDictPrd[ws.cell(i, 4).value] = ws.cell(i, 5).value
      else:
        orderDictPrd[ws.cell(i, 4).value] += ws.cell(i, 5).value
  
  for i in range(first_row, last_row):        
    if ws.cell(i, 7).value == None or ws.cell(i, 7).value == '':
      continue
    else:
      if ws.cell(i, 7).value not in orderDictSize:
        orderDictSize[ws.cell(i, 7).value] = ws.cell(i, 8).value
      else:
        orderDictSize[ws.cell(i, 7).value] += ws.cell(i, 8).value
        
  for i in range(first_row, last_row):        
    if ws.cell(i, 10).value == None or ws.cell(i, 10).value == '':
      continue
    else:
      if ws.cell(i, 10).value not in orderDictChannel:
        orderDictChannel[ws.cell(i, 10).value] = ws.cell(i, 11).value
      else:
        orderDictChannel[ws.cell(i, 10).value] += ws.cell(i, 11).value
        
  for i in range(first_row, last_row):        
    if ws.cell(i, 13).value == None or ws.cell(i, 13).value == '':
      continue
    else:
      if ws.cell(i, 13).value not in orderDictAddress:
        orderDictAddress[ws.cell(i, 13).value] = ws.cell(i, 14).value
      else:
        orderDictAddress[ws.cell(i, 13).value] += ws.cell(i, 14).value
        
  for i in range(first_row, last_row):        
    if ws.cell(i, 16).value == None or ws.cell(i, 16).value == '':
      continue
    else:
      if ws.cell(i, 16).value not in orderDictCustomer:
        orderDictCustomer[ws.cell(i, 16).value] = ws.cell(i, 17).value
      else:
        orderDictCustomer[ws.cell(i, 16).value] += ws.cell(i, 17).value
        
  # for i in range(first_row, last_row):        
  #   if ws.cell(i, 19).value == None or ws.cell(i, 19).value == '':
  #     continue
  #   else:
  #     if ws.cell(i, 19).value not in orderDictQuantity:
  #       orderDictQuantity[ws.cell(i, 19).value] = ws.cell(i, 20).value
  #     else:
  #       orderDictQuantity[ws.cell(i, 19).value] += ws.cell(i, 20).value
          
  for i in range(first_row, last_row):        
    if ws.cell(i, 22).value == None or ws.cell(i, 22).value == '':
      continue
    else:
      if ws.cell(i, 22).value not in orderDictPrdChannel:
        orderDictPrdChannel[ws.cell(i, 22).value] = ws.cell(i, 23).value
      else:
        orderDictPrdChannel[ws.cell(i, 22).value] += "/" + ws.cell(i, 23).value
          
  for i in range(first_row, last_row):        
    if ws.cell(i, 25).value == None or ws.cell(i, 25).value == '':
      continue
    else:
      if ws.cell(i, 25).value not in orderDictPrdDetailChannel:
        orderDictPrdDetailChannel[ws.cell(i, 25).value] = ws.cell(i, 26).value
      else:
        orderDictPrdDetailChannel[ws.cell(i, 25).value] += "/" + ws.cell(i, 26).value
          
  for i in range(first_row, last_row):        
    if ws.cell(i, 28).value == None or ws.cell(i, 28).value == '':
      continue
    else:
      if ws.cell(i, 28).value not in orderDictPrdNums:
        orderDictPrdNums[ws.cell(i, 28).value] = ws.cell(i, 29).value
      else:
        orderDictPrdNums[ws.cell(i, 28).value] += ws.cell(i, 29).value
          
  for i in range(first_row, last_row):        
    if ws.cell(i, 31).value == None or ws.cell(i, 31).value == '':
      continue
    else:
      if ws.cell(i, 31).value not in orderDictPrdDetailNums:
        orderDictPrdDetailNums[ws.cell(i, 31).value] = ws.cell(i, 32).value
      else:
        orderDictPrdDetailNums[ws.cell(i, 31).value] += ws.cell(i, 32).value
  
  orderDictCnt = 2
  for key, value in orderDict.items():
    accWs.cell(orderDictCnt, 1).value = key
    accWs.cell(orderDictCnt, 2).value = value
    orderDictCnt += 1
  
  orderDictPrdCnt = 2
  for key, value in orderDictPrd.items():
    accWs.cell(orderDictPrdCnt, 4).value = key
    accWs.cell(orderDictPrdCnt, 5).value = value
    orderDictPrdCnt += 1
  
  orderDictSizeCnt = 2
  for key, value in orderDictSize.items():
    accWs.cell(orderDictSizeCnt, 7).value = key
    accWs.cell(orderDictSizeCnt, 8).value = value
    orderDictSizeCnt += 1
  
  orderDictChannelCnt = 2
  for key, value in orderDictChannel.items():
    accWs.cell(orderDictChannelCnt, 10).value = key
    accWs.cell(orderDictChannelCnt, 11).value = value
    orderDictChannelCnt += 1
  
  orderDictAddressCnt = 2
  for key, value in orderDictAddress.items():
    accWs.cell(orderDictAddressCnt, 13).value = key
    accWs.cell(orderDictAddressCnt, 14).value = value
    orderDictAddressCnt += 1
  
  orderDictCustomerCnt = 2
  for key, value in orderDictCustomer.items():
    accWs.cell(orderDictCustomerCnt, 16).value = key
    accWs.cell(orderDictCustomerCnt, 17).value = value
    orderDictCustomerCnt += 1
  
  orderDictPrdChannelCnt = 2
  for key, value in orderDictPrdChannel.items():
    accWs.cell(orderDictPrdChannelCnt, 22).value = key
    accWs.cell(orderDictPrdChannelCnt, 23).value = "/".join(list(set(value.split('/'))))
    orderDictPrdChannelCnt += 1
  
  orderDictPrdDetailChannelCnt = 2
  for key, value in orderDictPrdDetailChannel.items():
    accWs.cell(orderDictPrdDetailChannelCnt, 25).value = key
    accWs.cell(orderDictPrdDetailChannelCnt, 26).value = "/".join(list(set(value.split('/'))))
    orderDictPrdDetailChannelCnt += 1
  
  orderDictPrdNumsCnt = 2
  for key, value in orderDictPrdNums.items():
    accWs.cell(orderDictPrdNumsCnt, 28).value = key
    accWs.cell(orderDictPrdNumsCnt, 29).value = value
    orderDictPrdNumsCnt += 1
  
  orderDictPrdDetailNumsCnt = 2
  for key, value in orderDictPrdDetailNums.items():
    accWs.cell(orderDictPrdDetailNumsCnt, 31).value = key
    accWs.cell(orderDictPrdDetailNumsCnt, 32).value = value
    orderDictPrdDetailNumsCnt += 1
    
# 주문수량기준 주문건수 정리
accWs = dailyWb.active

first_row = 2
last_row = accWs.max_row + 1

for i in range(first_row, last_row):
  try:
    if accWs.cell(i, 17).value >= 100:
      orderDictQuantity['100개 이상'] += 1
    elif accWs.cell(i, 17).value >= 50:
      orderDictQuantity['50개 이상'] += 1
    elif accWs.cell(i, 17).value >= 30:
      orderDictQuantity['30개 이상'] += 1
    elif accWs.cell(i, 17).value >= 20:
      orderDictQuantity['20개 이상'] += 1
    elif accWs.cell(i, 17).value >= 15:
      orderDictQuantity['15개 이상'] += 1
    elif accWs.cell(i, 17).value >= 10:
      orderDictQuantity['10개 이상'] += 1
    elif accWs.cell(i, 17).value >= 5:
      orderDictQuantity['5개 이상'] += 1
    elif accWs.cell(i, 17).value == 4:
      orderDictQuantity['4개'] += 1
    elif accWs.cell(i, 17).value == 3:
      orderDictQuantity['3개'] += 1
    elif accWs.cell(i, 17).value == 2:
      orderDictQuantity['2개'] += 1
    elif accWs.cell(i, 17).value == 1:
      orderDictQuantity['1개'] += 1
  except:
    pass
    
orderDictQuantityCnt = 2
for key, value in orderDictQuantity.items():
  accWs.cell(orderDictQuantityCnt, 19).value = key
  accWs.cell(orderDictQuantityCnt, 20).value = value
  print(accWs.cell(orderDictQuantityCnt, 19).value, accWs.cell(orderDictQuantityCnt, 20).value)
  orderDictQuantityCnt += 1

dailyWb.save(currPath + '\\' + fileName + '.xlsx')

import win32com.client

excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
wb = excel.Workbooks.Open(currPath + '\\' + fileName + '.xlsx')
ws = wb.Worksheets('판매량 체크')

ws.Range('A:B').Sort(Key1=ws.Range('B1'), Order1=2)
ws.Range('D:E').Sort(Key1=ws.Range('E1'), Order1=2)
ws.Range('G:H').Sort(Key1=ws.Range('H1'), Order1=2)
ws.Range('J:K').Sort(Key1=ws.Range('K1'), Order1=2)
ws.Range('M:N').Sort(Key1=ws.Range('N1'), Order1=2)
ws.Range('P:Q').Sort(Key1=ws.Range('Q1'), Order1=2)

# 주문수량기준은 판매량에 따른 정렬 미적용
# ws.Range('S:T').Sort(Key1=ws.Range('T1'), Order1=2)

wb.Save()

excel.Quit()