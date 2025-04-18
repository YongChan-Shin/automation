from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import os
from os import listdir
from os.path import exists
from os import makedirs

# 상품정보 리스트
product_list = []
color_list = []
size_list = []
cap_list = []

# DB 불러오기
import sqlite3
con = sqlite3.connect('D:/1.업무/10.기타자료/Development/db/productsData.db')
cur = con.cursor()

cur.execute("SELECT PrdName from ProductsData WHERE PrdName IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  product_list.append(i[0])

cur.execute("SELECT Color from ProductsData WHERE Color IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  color_list.append(i[0])

cur.execute("SELECT Size from ProductsData WHERE Size IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  size_list.append(i[0])

cur.execute("SELECT Cap from ProductsData WHERE Cap IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  cap_list.append(i[0])
  
# 폴더 내 엑셀 파일 검색
currPath = os.getcwd()
files = listdir(currPath + '\\data')
excelFileList = []

# 스마일비니 단품 판매건 리스트 
orderBeanie = []

for i in files:
  if(i.split('.')[-1] == 'xlsx'):
    if(not i.startswith('~')):
      excelFileList.append(i)

print(excelFileList)

# 채널명 정제 함수
def arrangeChannel(channel):
  if channel == "jja6806(옥션)":
    return "옥션"
  elif channel == "jja6806(지마켓)":
    return "지마켓"
  elif channel == "카카오스타일(저스틴23)":
    return "카카오스타일"
  elif channel == "위메프(저스틴23)":
    return "위메프"
  elif channel == "11번가(jja6806)":
    return "11번가"
  elif channel == "티몬(저스틴23)":
    return "티몬"
  elif channel == "스마트스토어(저스틴23)":
    return "스마트스토어"
  elif channel == "톡스토어(저스틴23)":
    return "톡스토어"
  elif channel == "톡스토어(저스틴23)_계정미사용":
    return "톡스토어"
  elif channel == "하프클럽":
    return "보리보리"
  elif channel == "키즈노트(저스틴23)":
    return "키즈노트"
  elif channel == "네오스토어_이몰":
    return "이몰"
  else:
    return channel

for file in excelFileList:

  wb = load_workbook(currPath + '\\data\\' + file)
  wb.create_sheet('추출내용')

  ws_name = wb.get_sheet_names()

  sheet1 = wb[str(ws_name[0])]
  sheet2 = wb[str(ws_name[1])]

  # ws = wb.active
  # sheetName = ws.title
  
  wb.active = wb['추출내용']

  for sheet in wb:
    if sheet.title == '추출내용':
      sheet.sheet_view.tabSelected = True
    else:
      sheet.sheet_view.tabSelected = False

  first_row = 2
  last_row = sheet1.max_row + 1
  first_col = 14
  last_col = sheet1.max_column + 1
  
  # 주문번호 수집
  orderDictPrdNums = {}
  orderDictPrdDetailNums = {}
  
  # 판매처 중복오류 체크
  salesChannelList = []
  baseChannel = ''

  for i in range(first_row, last_row):
    for j in range(first_col, last_col):
      sheet2MaxRow = sheet2.max_row
      if sheet1.cell(row=i, column=j).value == None or sheet1.cell(row=i, column=j).value == '':
        continue
      try:
        if sheet1.cell(row=i, column=7).value != baseChannel:
          baseChannel = sheet1.cell(row=i, column=7).value
          salesChannelList.append(baseChannel)
        print(sheet1.cell(row=i, column=7).value + " / 송장번호 : " + sheet1.cell(row=i, column=13).value)
      except:
        pass
      sheet2.cell(row=sheet2MaxRow + 1, column=1).value = sheet1.cell(row=i, column=j).value # 주문 정보 삽입
      sheet2.cell(row=sheet2MaxRow + 1, column=7).value = "_".join(sheet1.cell(i, 6).value.split(" ")[:2]) # 주소 정보 삽입
      sheet2.cell(row=sheet2MaxRow + 1, column=8).value = str(sheet1.cell(i, 4).value) + str(sheet1.cell(i, 3).value) # 주문 고객 정보 삽입
      
      # 상품(상세정보) 주문번호 수집 식별자
      prdDetailInfo = ''
      
      for product in product_list:
        try:
          if product in str(sheet1.cell(row=i, column=j).value):
            prdDetailInfoProduct = product.replace("(저스틴23)", "").replace("토밍이세트", "토밍이모자세트").replace("해피스노우세트", "해피스노우모자세트")
            if prdDetailInfoProduct not in orderDictPrdNums:
              orderDictPrdNums[prdDetailInfoProduct] = [sheet1.cell(row=i, column=9).value]
            else:
              if sheet1.cell(row=i, column=9).value not in orderDictPrdNums[prdDetailInfoProduct]:
                orderDictPrdNums[prdDetailInfoProduct].append(sheet1.cell(row=i, column=9).value) # 주문번호 정보 삽입
        except:
          pass
      
      for color in color_list:
        if color in str(sheet1.cell(row=i, column=j).value):
          prdDetailInfoColor = color
      for size in size_list:
        if size in str(sheet1.cell(row=i, column=j).value):
          prdDetailInfoSize = size.replace("FREE", "free")
      
      if prdDetailInfoProduct in cap_list:
        prdDetailInfoSize = "free"
      
      prdDetailInfo = '{}/{}/{}'.format(prdDetailInfoProduct, prdDetailInfoColor, prdDetailInfoSize)

      if prdDetailInfo not in orderDictPrdDetailNums:
        orderDictPrdDetailNums[prdDetailInfo] = [sheet1.cell(row=i, column=9).value]
      else:
        orderDictPrdDetailNums[prdDetailInfo].append(sheet1.cell(row=i, column=9).value)
      
  fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
  fillAlignment = Alignment(horizontal='center')
  fillFont = Font(bold=True)
  
  sheet2.cell(1, 1).value = '주문정보'
  sheet2.cell(1, 2).value = '상품명'
  sheet2.cell(1, 3).value = '컬러'
  sheet2.cell(1, 4).value = '사이즈'
  sheet2.cell(1, 5).value = '주문정보 정제'
  sheet2.cell(1, 6).value = '판매채널'
  sheet2.cell(1, 7).value = '주소'
  sheet2.cell(1, 8).value = '주문고객'
  sheet2.cell(1, 9).value = '주문건수(종합)'
  sheet2.cell(1, 10).value = '판매량'
  sheet2.cell(1, 12).value = '주문건수(상품기준)'
  sheet2.cell(1, 13).value = '판매량'
  sheet2.cell(1, 15).value = '주문건수(사이즈기준)'
  sheet2.cell(1, 16).value = '판매량'
  sheet2.cell(1, 18).value = '주문건수(채널기준)'
  sheet2.cell(1, 19).value = '판매량'
  sheet2.cell(1, 21).value = '주문건수(주소기준)'
  sheet2.cell(1, 22).value = '판매량'
  sheet2.cell(1, 24).value = '주문건수(주문고객기준)'
  sheet2.cell(1, 25).value = '판매량'
  sheet2.cell(1, 27).value = '주문건수(주문수량기준)'
  sheet2.cell(1, 28).value = '판매량'
  sheet2.cell(1, 30).value = '상품명'
  sheet2.cell(1, 31).value = '주문번호'
  sheet2.cell(1, 33).value = '상품명(상세)'
  sheet2.cell(1, 34).value = '주문번호'
  sheet2.cell(1, 36).value = '주문건수(사이즈기준 - 채널별 취합용)'
  sheet2.cell(1, 37).value = '판매량'
  sheet2.cell(1, 39).value = '주문건수(주문수량기준 - 채널별 취합용)'
  sheet2.cell(1, 40).value = '판매량'
  
  
  sheet2.cell(1, 1).alignment = fillAlignment
  sheet2.cell(1, 2).alignment = fillAlignment
  sheet2.cell(1, 3).alignment = fillAlignment
  sheet2.cell(1, 4).alignment = fillAlignment
  sheet2.cell(1, 5).alignment = fillAlignment
  sheet2.cell(1, 6).alignment = fillAlignment
  sheet2.cell(1, 7).alignment = fillAlignment
  sheet2.cell(1, 8).alignment = fillAlignment
  sheet2.cell(1, 9).alignment = fillAlignment
  sheet2.cell(1, 10).alignment = fillAlignment
  sheet2.cell(1, 12).alignment = fillAlignment
  sheet2.cell(1, 13).alignment = fillAlignment
  sheet2.cell(1, 15).alignment = fillAlignment
  sheet2.cell(1, 16).alignment = fillAlignment
  sheet2.cell(1, 18).alignment = fillAlignment
  sheet2.cell(1, 19).alignment = fillAlignment
  sheet2.cell(1, 21).alignment = fillAlignment
  sheet2.cell(1, 22).alignment = fillAlignment
  sheet2.cell(1, 24).alignment = fillAlignment
  sheet2.cell(1, 25).alignment = fillAlignment
  sheet2.cell(1, 27).alignment = fillAlignment
  sheet2.cell(1, 28).alignment = fillAlignment
  sheet2.cell(1, 30).alignment = fillAlignment
  sheet2.cell(1, 31).alignment = fillAlignment
  sheet2.cell(1, 33).alignment = fillAlignment
  sheet2.cell(1, 34).alignment = fillAlignment
  sheet2.cell(1, 36).alignment = fillAlignment
  sheet2.cell(1, 37).alignment = fillAlignment
  sheet2.cell(1, 39).alignment = fillAlignment
  sheet2.cell(1, 40).alignment = fillAlignment
  
  sheet2.cell(1, 1).font = fillFont
  sheet2.cell(1, 2).font = fillFont
  sheet2.cell(1, 3).font = fillFont
  sheet2.cell(1, 4).font = fillFont
  sheet2.cell(1, 5).font = fillFont
  sheet2.cell(1, 6).font = fillFont
  sheet2.cell(1, 7).font = fillFont
  sheet2.cell(1, 8).font = fillFont
  sheet2.cell(1, 9).font = fillFont
  sheet2.cell(1, 10).font = fillFont
  sheet2.cell(1, 12).font = fillFont
  sheet2.cell(1, 13).font = fillFont
  sheet2.cell(1, 15).font = fillFont
  sheet2.cell(1, 16).font = fillFont
  sheet2.cell(1, 18).font = fillFont
  sheet2.cell(1, 19).font = fillFont
  sheet2.cell(1, 21).font = fillFont
  sheet2.cell(1, 22).font = fillFont
  sheet2.cell(1, 24).font = fillFont
  sheet2.cell(1, 25).font = fillFont
  sheet2.cell(1, 27).font = fillFont
  sheet2.cell(1, 28).font = fillFont
  sheet2.cell(1, 30).font = fillFont
  sheet2.cell(1, 31).font = fillFont
  sheet2.cell(1, 33).font = fillFont
  sheet2.cell(1, 34).font = fillFont
  sheet2.cell(1, 36).font = fillFont
  sheet2.cell(1, 37).font = fillFont
  sheet2.cell(1, 39).font = fillFont
  sheet2.cell(1, 40).font = fillFont

  sheet2.cell(1, 1).fill = fillData
  sheet2.column_dimensions['A'].width = 60
  sheet2.column_dimensions['B'].width = 25
  sheet2.column_dimensions['C'].width = 10
  sheet2.column_dimensions['D'].width = 10
  sheet2.column_dimensions['E'].width = 40
  sheet2.column_dimensions['F'].width = 10
  sheet2.column_dimensions['G'].width = 20
  sheet2.column_dimensions['I'].width = 40
  sheet2.column_dimensions['L'].width = 40
  sheet2.column_dimensions['O'].width = 40
  sheet2.column_dimensions['R'].width = 40
  sheet2.column_dimensions['U'].width = 40
  sheet2.column_dimensions['X'].width = 40
  sheet2.column_dimensions['AA'].width = 40
  sheet2.column_dimensions['AD'].width = 40
  sheet2.column_dimensions['AE'].width = 40
  sheet2.column_dimensions['AG'].width = 40
  sheet2.column_dimensions['AH'].width = 40
  sheet2.column_dimensions['AJ'].width = 40
  sheet2.column_dimensions['AM'].width = 40

  last_row2 = sheet2.max_row + 1

  orderDict = {}
  orderDictPrd = {}
  orderDictSize = {}
  orderDictSizeChannelAcc = {}
  orderDictChannel = {}
  orderDictAddress = {}
  orderDictCustomer = {}
  orderDictQuantity = {
    '1개': 0,
    '2개': 0,
    '3개': 0,
    '4개': 0,
    '5개 이상': 0,
    '10개 이상': 0,
    '15개 이상': 0,
    '20개 이상': 0,
    '30개 이상': 0,
    '50개 이상': 0,
    '100개 이상': 0,
  }

  # 주문수량

  for i in range(first_row, last_row2):
    orderNum = sheet2.cell(i, 1).value.split('☞')[-1].replace("개", "")
    try:
      for product in product_list:
        if product in str(sheet2.cell(i, 1).value):
          sheet2.cell(i, 2).value = product.replace("(저스틴23)", "").replace("토밍이세트", "토밍이모자세트").replace("해피스노우세트", "해피스노우모자세트")
          # 스마일비니 단품 판매건 체크
          if "스마일비니" in str(sheet2.cell(i, 1).value) and "(1+1)" not in str(sheet2.cell(i, 1).value):
            orderBeanie.append(str(sheet1.cell(2, 7).value) + " / " + str(sheet2.cell(i, 1).value) + " / " + orderNum)
            
      for color in color_list:
        if color in str(sheet2.cell(i, 1).value):
          sheet2.cell(i, 3).value = color
      for size in size_list:
        if size in str(sheet2.cell(i, 1).value):
          sheet2.cell(i, 4).value = size.replace("FREE", "free")
      
      if str(sheet2.cell(i, 2).value.replace("(저스틴23)", "")) in cap_list:
        sheet2.cell(i, 4).value = "free"
        
      sheet2.cell(i, 5).value = str(sheet2.cell(i, 2).value.replace("(저스틴23)", "")) + "/" + str(sheet2.cell(i, 3).value) + "/" + str(sheet2.cell(i, 4).value.replace("FREE", "free"))
      sheet2.cell(i, 6).value = arrangeChannel(sheet1.cell(2, 7).value)
      
      if sheet2.cell(i, 5).value not in orderDict:
        orderDict[sheet2.cell(i, 5).value] = int(orderNum)
      else:
        orderDict[sheet2.cell(i, 5).value] += int(orderNum)
        
      if sheet2.cell(i, 2).value.replace("(저스틴23)", "") not in orderDictPrd:
        orderDictPrd[sheet2.cell(i, 2).value.replace("(저스틴23)", "")] = int(orderNum)
      else:
        orderDictPrd[sheet2.cell(i, 2).value.replace("(저스틴23)", "")] += int(orderNum)
        
      if sheet2.cell(i, 4).value not in orderDictSize:
        orderDictSize[sheet2.cell(i, 4).value] = int(orderNum)
        orderDictSizeChannelAcc[str(arrangeChannel(sheet1.cell(row=2, column=7).value)) + '/' + str(sheet2.cell(i, 4).value)] = int(orderNum)
      else:
        orderDictSize[sheet2.cell(i, 4).value] += int(orderNum)
        orderDictSizeChannelAcc[str(arrangeChannel(sheet1.cell(row=2, column=7).value)) + '/' + str(sheet2.cell(i, 4).value)] += int(orderNum)
      
      if sheet2.cell(i, 6).value not in orderDictChannel:
        orderDictChannel[sheet2.cell(i, 6).value] = int(orderNum)
      else:
        orderDictChannel[sheet2.cell(i, 6).value] += int(orderNum)
      
      if sheet2.cell(i, 7).value not in orderDictAddress:
        orderDictAddress[sheet2.cell(i, 7).value] = int(orderNum)
      else:
        orderDictAddress[sheet2.cell(i, 7).value] += int(orderNum)
      
      if sheet2.cell(i, 8).value not in orderDictCustomer:
        orderDictCustomer[sheet2.cell(i, 8).value] = int(orderNum)
      else:
        orderDictCustomer[sheet2.cell(i, 8).value] += int(orderNum)
        
      if sheet2.cell(i, 2).value == None or sheet2.cell(i, 3).value == None or sheet2.cell(i, 5).value == None or sheet2.cell(i, 6).value == None:
        fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
        sheet2["A{}".format(i)].fill = fillData
        sheet2["B{}".format(i)].fill = fillData
        sheet2["C{}".format(i)].fill = fillData
        sheet2["D{}".format(i)].fill = fillData
        sheet2["E{}".format(i)].fill = fillData
        sheet2["F{}".format(i)].fill = fillData
        sheet2["G{}".format(i)].fill = fillData
        sheet2["H{}".format(i)].fill = fillData
    except Exception as e:
      f = open('error.txt', 'a')
      f.write("{} / {} / {} / {}".format(file, i, sheet2.cell(i, 1).value, str(e)) + '\n')
      f.close()
  
  orderDictCnt = 2
  for key, value in orderDict.items():
    sheet2.cell(orderDictCnt, 9).value = key
    sheet2.cell(orderDictCnt, 10).value = value
    if(key == None or 'None' in key):
      sheet2["I{}".format(orderDictCnt)].fill = fillData
      sheet2["J{}".format(orderDictCnt)].fill = fillData
    orderDictCnt += 1
  
  orderDictPrdCnt = 2
  for key, value in orderDictPrd.items():
    sheet2.cell(orderDictPrdCnt, 12).value = key
    sheet2.cell(orderDictPrdCnt, 13).value = value
    if(key == None or 'None' in key):
      sheet2["L{}".format(orderDictPrdCnt)].fill = fillData
      sheet2["M{}".format(orderDictPrdCnt)].fill = fillData
    orderDictPrdCnt += 1
  
  orderDictSizeCnt = 2
  for key, value in orderDictSize.items():
    sheet2.cell(orderDictSizeCnt, 15).value = key
    sheet2.cell(orderDictSizeCnt, 16).value = value
    if(key == None or 'None' in key):
      sheet2["O{}".format(orderDictSizeCnt)].fill = fillData
      sheet2["P{}".format(orderDictSizeCnt)].fill = fillData
    orderDictSizeCnt += 1
  
  orderDictChannelCnt = 2
  for key, value in orderDictChannel.items():
    sheet2.cell(orderDictChannelCnt, 18).value = key
    sheet2.cell(orderDictChannelCnt, 19).value = value
    if(key == None or 'None' in key):
      sheet2["R{}".format(orderDictChannelCnt)].fill = fillData
      sheet2["S{}".format(orderDictChannelCnt)].fill = fillData
    orderDictChannelCnt += 1
  
  orderDictAddressCnt = 2
  for key, value in orderDictAddress.items():
    sheet2.cell(orderDictAddressCnt, 21).value = key
    sheet2.cell(orderDictAddressCnt, 22).value = value
    if(key == None or 'None' in key):
      sheet2["U{}".format(orderDictAddressCnt)].fill = fillData
      sheet2["V{}".format(orderDictAddressCnt)].fill = fillData
    orderDictAddressCnt += 1
  
  orderDictCustomerCnt = 2
  for key, value in orderDictCustomer.items():
    sheet2.cell(orderDictCustomerCnt, 24).value = key
    sheet2.cell(orderDictCustomerCnt, 25).value = value
    
    if sheet2.cell(orderDictCustomerCnt, 25).value >= 100:
      orderDictQuantity['100개 이상'] += 1
    elif sheet2.cell(orderDictCustomerCnt, 25).value >= 50:
      orderDictQuantity['50개 이상'] += 1
    elif sheet2.cell(orderDictCustomerCnt, 25).value >= 30:
      orderDictQuantity['30개 이상'] += 1
    elif sheet2.cell(orderDictCustomerCnt, 25).value >= 20:
      orderDictQuantity['20개 이상'] += 1
    elif sheet2.cell(orderDictCustomerCnt, 25).value >= 15:
      orderDictQuantity['15개 이상'] += 1
    elif sheet2.cell(orderDictCustomerCnt, 25).value >= 10:
      orderDictQuantity['10개 이상'] += 1
    elif sheet2.cell(orderDictCustomerCnt, 25).value >= 5:
      orderDictQuantity['5개 이상'] += 1
    elif sheet2.cell(orderDictCustomerCnt, 25).value == 4:
      orderDictQuantity['4개'] += 1
    elif sheet2.cell(orderDictCustomerCnt, 25).value == 3:
      orderDictQuantity['3개'] += 1
    elif sheet2.cell(orderDictCustomerCnt, 25).value == 2:
      orderDictQuantity['2개'] += 1
    elif sheet2.cell(orderDictCustomerCnt, 25).value == 1:
      orderDictQuantity['1개'] += 1
    
    if(key == None or 'None' in key):
      sheet2["X{}".format(orderDictCustomerCnt)].fill = fillData
      sheet2["Y{}".format(orderDictCustomerCnt)].fill = fillData
    
    orderDictCustomerCnt += 1

  orderDictQuantityCnt = 2
  for key, value in orderDictQuantity.items():
    sheet2.cell(orderDictQuantityCnt, 27).value = key
    sheet2.cell(orderDictQuantityCnt, 39).value = str(arrangeChannel(sheet1.cell(row=2, column=7).value)) + '/' + key
    sheet2.cell(orderDictQuantityCnt, 28).value = value
    sheet2.cell(orderDictQuantityCnt, 40).value = value
    orderDictQuantityCnt += 1
    
  orderDictPrdNumsCnt = 2
  for key, value in orderDictPrdNums.items():
    sheet2.cell(orderDictPrdNumsCnt, 30).value = key
    sheet2.cell(orderDictPrdNumsCnt, 31).value = ", ".join(value)
    orderDictPrdNumsCnt += 1
    
  orderDictPrdDetailNumsCnt = 2
  for key, value in orderDictPrdDetailNums.items():
    sheet2.cell(orderDictPrdDetailNumsCnt, 33).value = key
    sheet2.cell(orderDictPrdDetailNumsCnt, 34).value = ", ".join(value)
    orderDictPrdDetailNumsCnt += 1
    
  orderDictSizeChannelAccCnt = 2
  for key, value in orderDictSizeChannelAcc.items():
    sheet2.cell(orderDictSizeChannelAccCnt, 36).value = key
    sheet2.cell(orderDictSizeChannelAccCnt, 37).value = value
    orderDictSizeChannelAccCnt += 1    


  fillData2 = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
  fillData3 = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
  sheet2["A1"].fill = fillData2
  sheet2["B1"].fill = fillData2
  sheet2["C1"].fill = fillData2
  sheet2["D1"].fill = fillData2
  sheet2["E1"].fill = fillData2
  sheet2["F1"].fill = fillData2
  sheet2["G1"].fill = fillData2
  sheet2["H1"].fill = fillData2
  sheet2["I1"].fill = fillData3
  sheet2["J1"].fill = fillData3
  sheet2["L1"].fill = fillData3
  sheet2["M1"].fill = fillData3
  sheet2["O1"].fill = fillData3
  sheet2["P1"].fill = fillData3
  sheet2["R1"].fill = fillData3
  sheet2["S1"].fill = fillData3
  sheet2["U1"].fill = fillData3
  sheet2["V1"].fill = fillData3
  sheet2["X1"].fill = fillData3
  sheet2["Y1"].fill = fillData3
  sheet2["AA1"].fill = fillData3
  sheet2["AB1"].fill = fillData3
  sheet2["AD1"].fill = fillData2
  sheet2["AE1"].fill = fillData2
  sheet2["AG1"].fill = fillData2
  sheet2["AH1"].fill = fillData2
  sheet2["AJ1"].fill = fillData3
  sheet2["AK1"].fill = fillData3
  sheet2["AM1"].fill = fillData3
  sheet2["AN1"].fill = fillData3
  
  wb.active = sheet2

  if not exists(currPath + '\\주문건 정리본'):
    makedirs(currPath + '\\주문건 정리본')

  savePath = currPath + '\\주문건 정리본\\'

  wb.save(savePath + file + '_주문건 정리본.xlsx')
  
  if len(orderBeanie) > 0:
    f = open('스마일비니 단품 판매건.txt', 'w')
    for i in orderBeanie:
      f.write('{}\n'.format(i))
    f.close()
    
  salesChannelList = list(set(salesChannelList))
  
  if len(salesChannelList) > 1:
    f2 = open('판매처 중복오류 확인 필요_{}.txt'.format(file), 'w')
    for i in salesChannelList:
      f2.write('{}\n'.format(i))
    f2.close()