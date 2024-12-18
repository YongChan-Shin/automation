from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import os
from os import listdir
from os.path import exists
from os import makedirs
import productsData

# 상품정보 리스트
product_list = productsData.product_list
color_list = productsData.color_list
size_list = productsData.size_list

# 모자정보 리스트
capProducts = [
  "퍼피캡모자",
  "퍼피모자",
  "카우보이별모자",
  "곰돌이모자",
  "용용이모자",
  "베어캡모자",
  "알파벳모자",
  "알파벳벙거지",
  "보스턴챙모자",
  "스마일비니",
  "보스턴비니",
  "티라노모자",
  "토밍이세트",
  "토밍이모자세트",
  "코코모자",
  "토끼요정모자",
  "해피스노우세트",
  "해피스노우모자세트",
  "동물친구모자",
  "리리모자",
  "카우모자",
  "콩이모자",
  "포근이모자",
  "도토리비니",
  "카라멜비니",
  "왕방울모자",
]

# 폴더 내 엑셀 파일 검색
currPath = os.getcwd()

wb = load_workbook(currPath + '\\data.xlsx')
wb.create_sheet('추출내용')

ws_name = wb.get_sheet_names()

sheet1 = wb[str(ws_name[0])]
sheet2 = wb[str(ws_name[1])]

sheet1.column_dimensions['M'].width = 30
sheet1.column_dimensions['N'].width = 15
sheet1.column_dimensions['O'].width = 8
sheet1.column_dimensions['P'].width = 40

fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True)

sheet1.cell(3, 13).value = '상품명'
sheet1.cell(3, 14).value = '컬러'
sheet1.cell(3, 15).value = '사이즈'
sheet1.cell(3, 16).value = '상품식별정보'

sheet1.cell(3, 13).fill = fillData
sheet1.cell(3, 14).fill = fillData
sheet1.cell(3, 15).fill = fillData
sheet1.cell(3, 16).fill = fillData

wb.active = wb['추출내용']
# wb.active = wb['data']

for sheet in wb:
  if sheet.title == '추출내용':
    sheet.sheet_view.tabSelected = True
  else:
    sheet.sheet_view.tabSelected = False

first_row = 4
last_row = sheet1.max_row + 1

# 상품(상세정보) 식별자
prdDetailInfo = ''

orderDict = {}
orderDictPrd = {}
orderDictSize = {}

for i in range(first_row, last_row):
  for product in product_list:
    try:
      if product in str(sheet1.cell(row=i, column=2).value):
        prdDetailInfoProduct = product.replace("(저스틴23)", "").replace("토밍이세트", "토밍이모자세트").replace("해피스노우세트", "해피스노우모자세트")
        sheet1.cell(i, 13).value = prdDetailInfoProduct
        # if prdDetailInfoProduct not in orderDictPrd:
        #   orderDictPrd[prdDetailInfoProduct] = {'exposure': sheet1.cell(i, 4).value, 'click': sheet1.cell(i, 5).value, 'order': sheet1.cell(i, 6).value}
        # else:
        #   orderDictPrd[prdDetailInfoProduct]['exposure'] += sheet1.cell(i, 4).value
        #   orderDictPrd[prdDetailInfoProduct]['click'] += sheet1.cell(i, 5).value
        #   orderDictPrd[prdDetailInfoProduct]['order'] += sheet1.cell(i, 6).value
      for color in color_list:
        if color in str(sheet1.cell(row=i, column=2).value):
          prdDetailInfoColor = color
          sheet1.cell(i, 14).value = prdDetailInfoColor
      for size in size_list:
        if size in str(sheet1.cell(row=i, column=2).value):
          prdDetailInfoSize = size.replace("FREE", "free")
          sheet1.cell(i, 15).value = prdDetailInfoSize

      if prdDetailInfoProduct in capProducts:
        prdDetailInfoSize = "free"
        
      prdDetailInfo = '{}/{}/{}'.format(prdDetailInfoProduct, prdDetailInfoColor, prdDetailInfoSize)
      sheet1.cell(i, 16).value = prdDetailInfo
    
    except:
      pass
    
# sumExposure = 0
# sumClick = 0
# sumOrderCnt = 0

# orderDictPrdIdx = 4
# for key, value in orderDictPrd.items():
#   sumExposure += value['exposure']
#   sumClick += value['click']
#   sumOrderCnt += value['order']
#   sheet1.cell(orderDictPrdIdx, 17).value = key
#   sheet1.cell(orderDictPrdIdx, 18).value = value['exposure']
#   sheet1.cell(orderDictPrdIdx, 19).value = value['click']
#   sheet1.cell(orderDictPrdIdx, 20).value = value['order']
#   orderDictPrdIdx += 1
  
#   print(key, value)
  
# print(sumExposure, sumClick, sumOrderCnt)
  
      

# sheet2.cell(1, 1).value = '주문건수(종합)'
# sheet2.cell(1, 2).value = '판매량'
# sheet2.cell(1, 4).value = '주문건수(상품기준)'
# sheet2.cell(1, 5).value = '판매량'
# sheet2.cell(1, 7).value = '주문건수(사이즈기준)'
# sheet2.cell(1, 8).value = '판매량'

# sheet2.cell(1, 1).alignment = fillAlignment
# sheet2.cell(1, 2).alignment = fillAlignment
# sheet2.cell(1, 4).alignment = fillAlignment
# sheet2.cell(1, 5).alignment = fillAlignment
# sheet2.cell(1, 7).alignment = fillAlignment
# sheet2.cell(1, 8).alignment = fillAlignment

# sheet2.cell(1, 1).font = fillFont
# sheet2.cell(1, 2).font = fillFont
# sheet2.cell(1, 4).font = fillFont
# sheet2.cell(1, 5).font = fillFont
# sheet2.cell(1, 7).font = fillFont
# sheet2.cell(1, 8).font = fillFont

# sheet2.cell(1, 1).fill = fillData
# sheet2.cell(1, 2).fill = fillData
# sheet2.cell(1, 4).fill = fillData
# sheet2.cell(1, 5).fill = fillData
# sheet2.cell(1, 7).fill = fillData
# sheet2.cell(1, 8).fill = fillData

# sheet2.column_dimensions['A'].width = 40
# sheet2.column_dimensions['B'].width = 20
# sheet2.column_dimensions['D'].width = 40
# sheet2.column_dimensions['E'].width = 20
# sheet2.column_dimensions['G'].width = 40
# sheet2.column_dimensions['H'].width = 20

# last_row2 = sheet2.max_row + 1

# 주문수량

# for i in range(first_row, last_row2):
#   try:
#     for product in product_list:
#       if product in str(sheet2.cell(i, 1).value):
#         sheet2.cell(i, 2).value = product.replace("(저스틴23)", "").replace("토밍이세트", "토밍이모자세트").replace("해피스노우세트", "해피스노우모자세트")
          
#     for color in color_list:
#       if color in str(sheet2.cell(i, 1).value):
#         sheet2.cell(i, 3).value = color
#     for size in size_list:
#       if size in str(sheet2.cell(i, 1).value):
#         sheet2.cell(i, 4).value = size.replace("FREE", "free")
    
#     if str(sheet2.cell(i, 2).value.replace("(저스틴23)", "")) in capProducts:
#       sheet2.cell(i, 4).value = "free"
      
#     sheet2.cell(i, 5).value = str(sheet2.cell(i, 2).value.replace("(저스틴23)", "")) + "/" + str(sheet2.cell(i, 3).value) + "/" + str(sheet2.cell(i, 4).value.replace("FREE", "free"))
#     sheet2.cell(i, 6).value = sheet1.cell(2, 7).value
    
#     if sheet2.cell(i, 5).value not in orderDict:
#       orderDict[sheet2.cell(i, 5).value] = int(orderNum)
#     else:
#       orderDict[sheet2.cell(i, 5).value] += int(orderNum)
      
#     if sheet2.cell(i, 2).value.replace("(저스틴23)", "") not in orderDictPrd:
#       orderDictPrd[sheet2.cell(i, 2).value.replace("(저스틴23)", "")] = int(orderNum)
#     else:
#       orderDictPrd[sheet2.cell(i, 2).value.replace("(저스틴23)", "")] += int(orderNum)
      
#     if sheet2.cell(i, 4).value not in orderDictSize:
#       orderDictSize[sheet2.cell(i, 4).value] = int(orderNum)
#     else:
#       orderDictSize[sheet2.cell(i, 4).value] += int(orderNum)
    
#     if sheet2.cell(i, 6).value not in orderDictChannel:
#       orderDictChannel[sheet2.cell(i, 6).value] = int(orderNum)
#     else:
#       orderDictChannel[sheet2.cell(i, 6).value] += int(orderNum)
    
#     if sheet2.cell(i, 7).value not in orderDictAddress:
#       orderDictAddress[sheet2.cell(i, 7).value] = int(orderNum)
#     else:
#       orderDictAddress[sheet2.cell(i, 7).value] += int(orderNum)
    
#     if sheet2.cell(i, 8).value not in orderDictCustomer:
#       orderDictCustomer[sheet2.cell(i, 8).value] = int(orderNum)
#     else:
#       orderDictCustomer[sheet2.cell(i, 8).value] += int(orderNum)
      
#     if sheet2.cell(i, 2).value == None or sheet2.cell(i, 3).value == None or sheet2.cell(i, 5).value == None or sheet2.cell(i, 6).value == None:
#       fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
#       sheet2["A{}".format(i)].fill = fillData
#       sheet2["B{}".format(i)].fill = fillData
#       sheet2["C{}".format(i)].fill = fillData
#       sheet2["D{}".format(i)].fill = fillData
#       sheet2["E{}".format(i)].fill = fillData
#       sheet2["F{}".format(i)].fill = fillData
#       sheet2["G{}".format(i)].fill = fillData
#       sheet2["H{}".format(i)].fill = fillData
#   except Exception as e:
#     f = open('error.txt', 'a')
#     f.write("{} / {} / {} / {}".format(file, i, sheet2.cell(i, 1).value, str(e)) + '\n')
#     f.close()

# orderDictCnt = 2
# for key, value in orderDict.items():
#   sheet2.cell(orderDictCnt, 9).value = key
#   sheet2.cell(orderDictCnt, 10).value = value
#   if(key == None or 'None' in key):
#     sheet2["I{}".format(orderDictCnt)].fill = fillData
#     sheet2["J{}".format(orderDictCnt)].fill = fillData
#   orderDictCnt += 1

# orderDictPrdCnt = 2
# for key, value in orderDictPrd.items():
#   sheet2.cell(orderDictPrdCnt, 12).value = key
#   sheet2.cell(orderDictPrdCnt, 13).value = value
#   if(key == None or 'None' in key):
#     sheet2["L{}".format(orderDictPrdCnt)].fill = fillData
#     sheet2["M{}".format(orderDictPrdCnt)].fill = fillData
#   orderDictPrdCnt += 1

# orderDictSizeCnt = 2
# for key, value in orderDictSize.items():
#   sheet2.cell(orderDictSizeCnt, 15).value = key
#   sheet2.cell(orderDictSizeCnt, 16).value = value
#   if(key == None or 'None' in key):
#     sheet2["O{}".format(orderDictSizeCnt)].fill = fillData
#     sheet2["P{}".format(orderDictSizeCnt)].fill = fillData
#   orderDictSizeCnt += 1

# orderDictChannelCnt = 2
# for key, value in orderDictChannel.items():
#   sheet2.cell(orderDictChannelCnt, 18).value = key
#   sheet2.cell(orderDictChannelCnt, 19).value = value
#   if(key == None or 'None' in key):
#     sheet2["R{}".format(orderDictChannelCnt)].fill = fillData
#     sheet2["S{}".format(orderDictChannelCnt)].fill = fillData
#   orderDictChannelCnt += 1

# orderDictAddressCnt = 2
# for key, value in orderDictAddress.items():
#   sheet2.cell(orderDictAddressCnt, 21).value = key
#   sheet2.cell(orderDictAddressCnt, 22).value = value
#   if(key == None or 'None' in key):
#     sheet2["U{}".format(orderDictAddressCnt)].fill = fillData
#     sheet2["V{}".format(orderDictAddressCnt)].fill = fillData
#   orderDictAddressCnt += 1

# orderDictCustomerCnt = 2
# for key, value in orderDictCustomer.items():
#   sheet2.cell(orderDictCustomerCnt, 24).value = key
#   sheet2.cell(orderDictCustomerCnt, 25).value = value
  
#   if sheet2.cell(orderDictCustomerCnt, 25).value >= 100:
#     orderDictQuantity['100개 이상'] += 1
#   elif sheet2.cell(orderDictCustomerCnt, 25).value >= 50:
#     orderDictQuantity['50개 이상'] += 1
#   elif sheet2.cell(orderDictCustomerCnt, 25).value >= 30:
#     orderDictQuantity['30개 이상'] += 1
#   elif sheet2.cell(orderDictCustomerCnt, 25).value >= 20:
#     orderDictQuantity['20개 이상'] += 1
#   elif sheet2.cell(orderDictCustomerCnt, 25).value >= 15:
#     orderDictQuantity['15개 이상'] += 1
#   elif sheet2.cell(orderDictCustomerCnt, 25).value >= 10:
#     orderDictQuantity['10개 이상'] += 1
#   elif sheet2.cell(orderDictCustomerCnt, 25).value >= 5:
#     orderDictQuantity['5개 이상'] += 1
#   elif sheet2.cell(orderDictCustomerCnt, 25).value == 4:
#     orderDictQuantity['4개'] += 1
#   elif sheet2.cell(orderDictCustomerCnt, 25).value == 3:
#     orderDictQuantity['3개'] += 1
#   elif sheet2.cell(orderDictCustomerCnt, 25).value == 2:
#     orderDictQuantity['2개'] += 1
#   elif sheet2.cell(orderDictCustomerCnt, 25).value == 1:
#     orderDictQuantity['1개'] += 1
  
#   if(key == None or 'None' in key):
#     sheet2["X{}".format(orderDictCustomerCnt)].fill = fillData
#     sheet2["Y{}".format(orderDictCustomerCnt)].fill = fillData
  
#   orderDictCustomerCnt += 1

# orderDictQuantityCnt = 2
# for key, value in orderDictQuantity.items():
#   sheet2.cell(orderDictQuantityCnt, 27).value = key
#   sheet2.cell(orderDictQuantityCnt, 28).value = value
#   orderDictQuantityCnt += 1
  
# orderDictPrdNumsCnt = 2
# for key, value in orderDictPrdNums.items():
#   sheet2.cell(orderDictPrdNumsCnt, 30).value = key
#   sheet2.cell(orderDictPrdNumsCnt, 31).value = ", ".join(value)
#   orderDictPrdNumsCnt += 1
  
# orderDictPrdDetailNumsCnt = 2
# for key, value in orderDictPrdDetailNums.items():
#   sheet2.cell(orderDictPrdDetailNumsCnt, 33).value = key
#   sheet2.cell(orderDictPrdDetailNumsCnt, 34).value = ", ".join(value)
#   orderDictPrdDetailNumsCnt += 1


# fillData2 = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
# fillData3 = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
# sheet2["A1"].fill = fillData2
# sheet2["B1"].fill = fillData2
# sheet2["C1"].fill = fillData2
# sheet2["D1"].fill = fillData2
# sheet2["E1"].fill = fillData2
# sheet2["F1"].fill = fillData2
# sheet2["G1"].fill = fillData2
# sheet2["H1"].fill = fillData2
# sheet2["I1"].fill = fillData3
# sheet2["J1"].fill = fillData3
# sheet2["L1"].fill = fillData3
# sheet2["M1"].fill = fillData3
# sheet2["O1"].fill = fillData3
# sheet2["P1"].fill = fillData3
# sheet2["R1"].fill = fillData3
# sheet2["S1"].fill = fillData3
# sheet2["U1"].fill = fillData3
# sheet2["V1"].fill = fillData3
# sheet2["X1"].fill = fillData3
# sheet2["Y1"].fill = fillData3
# sheet2["AA1"].fill = fillData3
# sheet2["AB1"].fill = fillData3
# sheet2["AD1"].fill = fillData2
# sheet2["AE1"].fill = fillData2
# sheet2["AG1"].fill = fillData2
# sheet2["AH1"].fill = fillData2

wb.save(currPath + '\\data_정리본.xlsx')