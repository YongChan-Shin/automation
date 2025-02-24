from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.drawing.image import Image
import os
from datetime import datetime

wb = load_workbook('판매데이터.xlsx')
sheetName = '상품사진삽입'

currPath = os.getcwd()
now = datetime.now()

# 제품 코드정보 생성
productsCode = {}

# DB 불러오기
import sqlite3
con = sqlite3.connect('D:/1.업무/10.기타자료/Development/db/productsCode.db')
cur = con.cursor()

cur.execute("SELECT PrdName, PrdCode from ProductsCode")
data = cur.fetchall()

for i in data:
  productsCode[i[0]] = i[1]

firstCell = 2
lastCell = wb[sheetName].max_row + 1

# 판매정보 생성
sellList = {}
for i in range(firstCell, lastCell):
  if wb[sheetName].cell(row=i, column=1).value == None or wb[sheetName].cell(row=i, column=1).value == '':
    continue
  else:
    sellList[wb[sheetName].cell(i, 1).value.replace('(저스틴23)', '').replace('/', ' ')] = wb[sheetName].cell(i, 2).value

print(sellList)

fileName = input("판매량 순위 산출 기준 파일명 입력 : ")

wb2 = load_workbook(fileName + '.xlsx')
wb2Sheet = wb2.active
wb2FirstCell = 2
wb2LastCell = wb2Sheet.max_row + 1
wb2LastColumn = wb2Sheet.max_column + 1

fillData = PatternFill(fill_type='solid', start_color='000000', end_color='000000')
fillAlignment = Alignment(horizontal='center', vertical='center')
fillFont = Font(color='FFFFFF')

wb2Sheet.cell(1, wb2LastColumn).value = '판매량'
wb2Sheet.cell(1, wb2LastColumn + 1).value = '이미지'
wb2Sheet.cell(1, wb2LastColumn).fill = fillData
wb2Sheet.cell(1, wb2LastColumn + 1).fill = fillData
wb2Sheet.cell(1, wb2LastColumn).alignment = fillAlignment
wb2Sheet.cell(1, wb2LastColumn + 1).alignment = fillAlignment
wb2Sheet.cell(1, wb2LastColumn).font = fillFont
wb2Sheet.cell(1, wb2LastColumn + 1).font = fillFont

# TODO 기준 파일에 맞춰 정보 수정 필요(이미지 삽입 열 너비 지정)
wb2Sheet.column_dimensions['J'].width = 12.5

for i in range(wb2FirstCell, wb2LastCell):
  try:
    # TODO 기준 파일에 맞춰 정보 수정 필요
    wb2Sheet.cell(i, wb2LastColumn).value = sellList[wb2Sheet.cell(i, 4).value]
  except:
    wb2Sheet.cell(i, wb2LastColumn).value = 0
    
  try:
    wb2Sheet.row_dimensions[i].height = 75
    # TODO 기준 파일에 맞춰 정보 수정 필요
    image_path = '.\\data\\images\\' + str(productsCode[wb2Sheet.cell(row=i, column=4).value]) + '.jpg'
    print(image_path)
    image = Image(image_path)
    image.width = 100
    image.height = 100
    # TODO 기준 파일에 맞춰 정보 수정 필요(이미지 삽입 열 지정)
    wb2Sheet.add_image(image, anchor='J'+str(i))
  except:
    pass
    
wb2.save(currPath + '\\' + fileName + '.xlsx')

import win32com.client

excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
wb = excel.Workbooks.Open(currPath + '\\' + fileName + '.xlsx')

# TODO 기준 파일에 맞춰 정보 수정 필요
ws = wb.Worksheets('2025_봄')

# TODO 기준 파일에 맞춰 정보 수정 필요
ws.Range('A:J').Sort(Key1=ws.Range('I1'), Order1=2)

wb.Save()

excel.Quit()