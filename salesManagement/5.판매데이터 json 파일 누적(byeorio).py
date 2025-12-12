from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import os
from os import listdir
from os.path import exists
from os import makedirs
import math
import datetime
import json
import shutil

# json 파일 백업
from_jsonFile_path = './json(byeorio)/salesTotal.json'
to_jsonFile_path = './json(byeorio)/total_backup/salesTotal_backup.json'
shutil.copy(from_jsonFile_path, to_jsonFile_path)

# 폴더 내 파일 검색
currPath = os.getcwd()
files = listdir(currPath)
jsonFileList = []

# json 판매 데이터 누적
dailyTotalData = []

for i in files:
  if(i.split('.')[-1] == 'json'):
    if i != 'json':
      jsonFileList.append(i)

for file in jsonFileList:
  with open(file, encoding='UTF-8') as d:
    dailyData = json.load(d)
    dailyTotalData.append(dailyData['data'][0])
  
with open('./json(byeorio)/salesTotal.json', 'r', encoding='UTF-8') as t:
  salesTotalData = json.load(t)

with open('./json(byeorio)/salesTotal.json', 'w', encoding='UTF-8') as outfile:
  for data in dailyTotalData:
    salesTotalData['data'].append(data)
  json.dump(salesTotalData, outfile, indent=2, ensure_ascii=False)