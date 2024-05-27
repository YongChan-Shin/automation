from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

wb = load_workbook('data.xlsx')
sheet1Name = '상품리스트'
sheet2Name = '품절상품체크'

currPath = os.getcwd()

firstCell = 2
lastCell = wb[sheet1Name].max_row + 1

wb.active = wb[sheet2Name]

prdIdxInfoList = []
for row in wb[sheet1Name].iter_rows(min_row=firstCell, max_row=lastCell, min_col=6, max_col=6):
    for cell in row:
        prdIdxInfoList.append(cell.value)

fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')

for sheet in wb:
  if sheet.title == sheet2Name:
    sheet.sheet_view.tabSelected = True
  else:
    sheet.sheet_view.tabSelected = False

cnt = 2

for i in range(firstCell, lastCell):
  if wb[sheet1Name].cell(i, 3).value == None or wb[sheet1Name].cell(i, 3).value == '':
    continue
  colorList = wb[sheet1Name].cell(i, 3).value.split('/')
  for color in colorList:
    sizeList = wb[sheet1Name].cell(i, 4).value.split('/')
    for size in sizeList:
      for j in range(1, 6):
        wb[sheet2Name].cell(cnt, j).value = wb[sheet1Name].cell(i, j).value
      wb[sheet2Name].cell(cnt, 3).value = color
      wb[sheet2Name].cell(cnt, 4).value = size
      wb[sheet2Name].cell(cnt, 5).value = wb[sheet2Name].cell(cnt, 2).value + "/" + color + "/" + size
      prdIdxInfo = wb[sheet2Name].cell(cnt, 5).value
      
      if prdIdxInfo in prdIdxInfoList:
        wb[sheet2Name].cell(cnt, 1).fill = fillData
        wb[sheet2Name].cell(cnt, 2).fill = fillData
        wb[sheet2Name].cell(cnt, 3).fill = fillData
        wb[sheet2Name].cell(cnt, 4).fill = fillData
        wb[sheet2Name].cell(cnt, 5).fill = fillData
        wb[sheet2Name].cell(cnt, 6).fill = fillData
        wb[sheet2Name].cell(cnt, 6).value = "품절"
      
      cnt += 1
  

wb.save(currPath + '\\data_품절상품 체크.xlsx')