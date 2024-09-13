from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from shutil import copyfile
import os

copyfile('data.xlsx', '태그 자동생성 및 추출.xlsx')

wb = load_workbook('태그 자동생성 및 추출.xlsx')
ws = wb.active



first_cell = 3
last_cell = ws.max_row + 1
cell_bgcolor = PatternFill(start_color='0018FF', end_color='0018FF', fill_type='solid')
cnt = 1
source_text = f'''
<div align="center">
<table border="0"  cellspacing="0" cellpadding="0" align="center">
<tbody>
<img src="http://gi.esmplus.com/jst4994/data/notice.jpg">'''

for i in range(first_cell, last_cell):
  url_source = ws.cell(row=i, column=5).value
  url = requests.get(url_source)
  soup = BeautifulSoup(url.content, 'html.parser')
  
  mergedSource = ''
  if ws.cell(1, 3).value == '' or ws.cell(1, 3).value is None:
    mergedSource += str(soup)
  else:
    mergedSource += f'''<center>
<img src="{ws.cell(1, 3).value}" border="0">
</center>
''' + str(soup)
  if ws.cell(1, 4).value == '' or ws.cell(1, 4).value is None:
    pass
  else:
    mergedSource += f'''
<center>
<img src="{ws.cell(1, 4).value}" border="0">
</center>'''
  
  # ws.cell(row=i, column=6).value = str(soup)
  ws.cell(row=i, column=6).value = mergedSource
  num = ws.cell(row=i, column=1).value
  folder_name = ws.cell(row=1, column=2).value
  
  print(mergedSource)
  
  if num < 10:
    num = "0" + str(num)
    
  if cnt % 2 != 0:
    source_text += f'''
<tr>		
<td><a href="{url_source}" target="pop">
<img src="{folder_name}/list_{num}.jpg" border="0" ></a></td>'''
  else:
    source_text += f'''
<td><a href="{url_source}" target="pop">
<img src="{folder_name}/list_{num}.jpg" border="0" ></a></td>		
</tr>'''

  cnt += 1
  
source_text += f'''
</tbody>
</table>'''
  
ws.merge_cells(start_row=first_cell, start_column=7, end_row=last_cell-1, end_column=7)
ws.cell(row=2, column=6).fill = cell_bgcolor
ws.cell(row=2, column=6).font = Font(color='FFFFFF')
ws.cell(row=2, column=6).value = '태그추출'
ws.cell(row=2, column=7).fill = cell_bgcolor
ws.cell(row=2, column=7).font = Font(color='FFFFFF')
ws.cell(row=2, column=7).value = '소스생성'
ws.cell(row=3, column=7).value = source_text
  
wb.save('태그 자동생성 및 추출.xlsx')