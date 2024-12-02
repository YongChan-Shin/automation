from bs4 import BeautifulSoup
import requests
import urllib.request
import urllib
from PIL import Image
from openpyxl import load_workbook
import os

wb = load_workbook('data.xlsx')
ws = wb.active

first_cell = 2
last_cell = ws.max_row + 1
# last_cell = 4
cnt = 1

for num in range(first_cell, last_cell):
  
  url_source = ws.cell(row=num, column=5).value
  
  if url_source.split('.')[-1] == 'jpg':
    print(url_source)
    urllib.request.urlretrieve(url_source, '0.jpg')
    image1 = Image.open('0.jpg')
    new_image = Image.new('RGB', (image1.width, image1.height), color=(255, 255, 255))
    new_image.paste(image1, (0, 0))
    
    if new_image.width > 780:
      diff_width = new_image.width - 780
      diff_ratio = 1 - (diff_width / new_image.width)
    else:
      diff_ratio = 1
      
    img_resize = new_image.resize((int(new_image.width * diff_ratio), int(new_image.height * diff_ratio))) # 이미지 기본 너비 780px 이하로 축소    
    
    if cnt < 10:
      new_image.save('img/detail_0' + str(cnt) + '.jpg', 'JPEG', quality=100, subsampling=0)
      img_resize.save('img/detail_0' + str(cnt) + '.jpg', 'JPEG', quality=83, subsampling=0)
    else:
      new_image.save('img/detail_' + str(cnt) + '.jpg', 'JPEG', quality=100, subsampling=0)
      img_resize.save('img/detail_' + str(cnt) + '.jpg', 'JPEG', quality=83, subsampling=0)
      
    cnt += 1
      
    continue
  
  url = requests.get(url_source)
  soup = BeautifulSoup(url.content, 'html.parser')

  img_src = soup.find_all('img')

  img_list = []
  for i in img_src:
      img_src1 = i['src']
      img_list.append(img_src1)
      

  file_no = 0
  file_num = ws.cell(row=num, column=1).value
  file_name = ws.cell(row=num, column=3).value
  code_name = ws.cell(row=num, column=2).value
  

  for j in range(0, len(img_list)):
      try:
          urllib.request.urlretrieve(img_list[j], str(file_no) + '.jpg')
          file_no += 1
      except:
          continue
  
  white = (255, 255, 255)
  
  print(str(cnt) + ' / ' + str(img_list))
  
  if len(img_list) == 1:      
    image1 = Image.open('0.jpg')
    new_image = Image.new('RGB', (image1.width, image1.height), color=(255, 255, 255))
    new_image.paste(image1, (0, 0))
    
  if len(img_list) == 2:      
    image1 = Image.open('0.jpg')
    image2 = Image.open('1.jpg')
    
    max_width = max(image1.width, image2.width)
    min_width = min(image1.width, image2.width)
    center_width_1 = (max_width - image1.width) // 2
    center_width_2 = (max_width - image2.width) // 2
    
    new_image = Image.new('RGB', (max_width, image1.height + image2.height), color=(255, 255, 255))
    new_image.paste(image1, (center_width_1, 0))
    new_image.paste(image2, (center_width_2, image1.height))
    
  if len(img_list) == 3:      
    image1 = Image.open('0.jpg')
    image2 = Image.open('1.jpg')
    image3 = Image.open('2.jpg')
    
    max_width = max(image1.width, image2.width, image3.width)
    min_width = min(image1.width, image2.width, image3.width)
    center_width_1 = (max_width - image1.width) // 2
    center_width_2 = (max_width - image2.width) // 2
    center_width_3 = (max_width - image3.width) // 2
    
    new_image = Image.new('RGB', (max_width, image1.height + image2.height + image3.height), color=(255, 255, 255))
    new_image.paste(image1, (center_width_1, 0))
    new_image.paste(image2, (center_width_2, image1.height))
    new_image.paste(image3, (center_width_3, image1.height + image2.height))
    
  if len(img_list) == 4:
    image1 = Image.open('0.jpg')
    image2 = Image.open('1.jpg')
    image3 = Image.open('2.jpg')
    image4 = Image.open('3.jpg')
    
    max_width = max(image1.width, image2.width, image3.width, image4.width)
    min_width = min(image1.width, image2.width, image3.width, image4.width)
    center_width_1 = (max_width - image1.width) // 2
    center_width_2 = (max_width - image2.width) // 2
    center_width_3 = (max_width - image3.width) // 2
    center_width_4 = (max_width - image4.width) // 2
    
    new_image = Image.new('RGB', (max_width, image1.height + image2.height + image3.height + image4.height), color=(255, 255, 255))
    new_image.paste(image1, (center_width_1, 0))
    new_image.paste(image2, (center_width_2, image1.height))
    new_image.paste(image3, (center_width_3, image1.height + image2.height))
    new_image.paste(image4, (center_width_4, image1.height + image2.height + image3.height))

  if len(img_list) == 5:      
    image1 = Image.open('0.jpg')
    image2 = Image.open('1.jpg')
    image3 = Image.open('2.jpg')
    image4 = Image.open('3.jpg')
    image5 = Image.open('4.jpg')
    
    max_width = max(image1.width, image2.width, image3.width, image4.width, image5.width)
    min_width = min(image1.width, image2.width, image3.width, image4.width, image5.width)
    center_width_1 = (max_width - image1.width) // 2
    center_width_2 = (max_width - image2.width) // 2
    center_width_3 = (max_width - image3.width) // 2
    center_width_4 = (max_width - image4.width) // 2
    center_width_5 = (max_width - image5.width) // 2
    
    new_image = Image.new('RGB', (max_width, image1.height + image2.height + image3.height + image4.height + image5.height), color=(255, 255, 255))
    new_image.paste(image1, (center_width_1, 0))
    new_image.paste(image2, (center_width_2, image1.height))
    new_image.paste(image3, (center_width_3, image1.height + image2.height))
    new_image.paste(image4, (center_width_4, image1.height + image2.height + image3.height))
    new_image.paste(image5, (center_width_5, image1.height + image2.height + image3.height + image4.height))
    
  if len(img_list) == 6:      
    image1 = Image.open('0.jpg')
    image2 = Image.open('1.jpg')
    image3 = Image.open('2.jpg')
    image4 = Image.open('3.jpg')
    image5 = Image.open('4.jpg')
    image6 = Image.open('5.jpg')
    
    max_width = max(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width)
    min_width = min(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width)
    center_width_1 = (max_width - image1.width) // 2
    center_width_2 = (max_width - image2.width) // 2
    center_width_3 = (max_width - image3.width) // 2
    center_width_4 = (max_width - image4.width) // 2
    center_width_5 = (max_width - image5.width) // 2
    center_width_6 = (max_width - image6.width) // 2
    
    new_image = Image.new('RGB', (max_width, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height), color=(255, 255, 255))
    new_image.paste(image1, (center_width_1, 0))
    new_image.paste(image2, (center_width_2, image1.height))
    new_image.paste(image3, (center_width_3, image1.height + image2.height))
    new_image.paste(image4, (center_width_4, image1.height + image2.height + image3.height))
    new_image.paste(image5, (center_width_5, image1.height + image2.height + image3.height + image4.height))
    new_image.paste(image6, (center_width_6, image1.height + image2.height + image3.height + image4.height + image5.height))
    
  if len(img_list) == 7:      
    image1 = Image.open('0.jpg')
    image2 = Image.open('1.jpg')
    image3 = Image.open('2.jpg')
    image4 = Image.open('3.jpg')
    image5 = Image.open('4.jpg')
    image6 = Image.open('5.jpg')
    image7 = Image.open('6.jpg')
    
    max_width = max(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width, image7.width)
    min_width = min(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width, image7.width)
    center_width_1 = (max_width - image1.width) // 2
    center_width_2 = (max_width - image2.width) // 2
    center_width_3 = (max_width - image3.width) // 2
    center_width_4 = (max_width - image4.width) // 2
    center_width_5 = (max_width - image5.width) // 2
    center_width_6 = (max_width - image6.width) // 2
    center_width_7 = (max_width - image7.width) // 2
    
    new_image = Image.new('RGB', (max_width, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height), color=(255, 255, 255))
    new_image.paste(image1, (center_width_1, 0))
    new_image.paste(image2, (center_width_2, image1.height))
    new_image.paste(image3, (center_width_3, image1.height + image2.height))
    new_image.paste(image4, (center_width_4, image1.height + image2.height + image3.height))
    new_image.paste(image5, (center_width_5, image1.height + image2.height + image3.height + image4.height))
    new_image.paste(image6, (center_width_6, image1.height + image2.height + image3.height + image4.height + image5.height))
    new_image.paste(image7, (center_width_7, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height))
    
  if len(img_list) == 8:      
    image1 = Image.open('0.jpg')
    image2 = Image.open('1.jpg')
    image3 = Image.open('2.jpg')
    image4 = Image.open('3.jpg')
    image5 = Image.open('4.jpg')
    image6 = Image.open('5.jpg')
    image7 = Image.open('6.jpg')
    image8 = Image.open('7.jpg')
    
    # 최하단 공지 제외 크롤링 시
    # image8 = Image.open('7.jpg')
    
    max_width = max(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width, image7.width, image8.width)
    min_width = min(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width, image7.width, image8.width)
    
    # 최하단 공지 제외 크롤링 시
    # max_width = max(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width, image7.width)
    # min_width = min(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width, image7.width)
    
    center_width_1= (max_width - image1.width) // 2
    center_width_2= (max_width - image2.width) // 2
    center_width_3= (max_width - image3.width) // 2
    center_width_4= (max_width - image4.width) // 2
    center_width_5= (max_width - image5.width) // 2
    center_width_6= (max_width - image6.width) // 2
    center_width_7= (max_width - image7.width) // 2
    center_width_8= (max_width - image8.width) // 2
    
    new_image = Image.new('RGB', (max_width, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height + image8.height), color=(255, 255, 255))
    
    # 최하단 공지 제외 크롤링 시
    # new_image = Image.new('RGB', (max_width, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height), color=(255, 255, 255))
    
    new_image.paste(image1, (center_width_1, 0))
    new_image.paste(image2, (center_width_2, image1.height))
    new_image.paste(image3, (center_width_3, image1.height + image2.height))
    new_image.paste(image4, (center_width_4, image1.height + image2.height + image3.height))
    new_image.paste(image5, (center_width_5, image1.height + image2.height + image3.height + image4.height))
    new_image.paste(image6, (center_width_6, image1.height + image2.height + image3.height + image4.height + image5.height))
    new_image.paste(image7, (center_width_7, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height))
    new_image.paste(image8, (center_width_7, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height))
    
  if len(img_list) == 9:      
    image1 = Image.open('0.jpg')
    image2 = Image.open('1.jpg')
    image3 = Image.open('2.jpg')
    image4 = Image.open('3.jpg')
    image5 = Image.open('4.jpg')
    image6 = Image.open('5.jpg')
    image7 = Image.open('6.jpg')
    image8 = Image.open('7.jpg')
    image9 = Image.open('8.jpg')
    
    max_width = max(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width, image7.width, image8.width, image9.width)
    min_width = min(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width, image7.width, image8.width, image9.width)
    center_width_1= (max_width - image1.width) // 2
    center_width_2= (max_width - image2.width) // 2
    center_width_3= (max_width - image3.width) // 2
    center_width_4= (max_width - image4.width) // 2
    center_width_5= (max_width - image5.width) // 2
    center_width_6= (max_width - image6.width) // 2
    center_width_7= (max_width - image7.width) // 2
    center_width_8= (max_width - image8.width) // 2
    center_width_9= (max_width - image9.width) // 2
    
    new_image = Image.new('RGB', (max_width, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height + image8.height + image9.height), color=(255, 255, 255))
    new_image.paste(image1, (center_width_1, 0))
    new_image.paste(image2, (center_width_2, image1.height))
    new_image.paste(image3, (center_width_3, image1.height + image2.height))
    new_image.paste(image4, (center_width_4, image1.height + image2.height + image3.height))
    new_image.paste(image5, (center_width_5, image1.height + image2.height + image3.height + image4.height))
    new_image.paste(image6, (center_width_6, image1.height + image2.height + image3.height + image4.height + image5.height))
    new_image.paste(image7, (center_width_7, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height))
    new_image.paste(image8, (center_width_7, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height))
    new_image.paste(image9, (center_width_7, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height + image8.height))
    
  if len(img_list) == 10:      
    image1 = Image.open('0.jpg')
    image2 = Image.open('1.jpg')
    image3 = Image.open('2.jpg')
    image4 = Image.open('3.jpg')
    image5 = Image.open('4.jpg')
    image6 = Image.open('5.jpg')
    image7 = Image.open('6.jpg')
    image8 = Image.open('7.jpg')
    image9 = Image.open('8.jpg')
    image10 = Image.open('9.jpg')
    
    max_width = max(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width, image7.width, image8.width, image9.width, image10.width)
    min_width = min(image1.width, image2.width, image3.width, image4.width, image5.width, image6.width, image7.width, image8.width, image9.width, image10.width)
    center_width_1= (max_width - image1.width) // 2
    center_width_2= (max_width - image2.width) // 2
    center_width_3= (max_width - image3.width) // 2
    center_width_4= (max_width - image4.width) // 2
    center_width_5= (max_width - image5.width) // 2
    center_width_6= (max_width - image6.width) // 2
    center_width_7= (max_width - image7.width) // 2
    center_width_8= (max_width - image8.width) // 2
    center_width_9= (max_width - image9.width) // 2
    center_width_10= (max_width - image10.width) // 2
    
    new_image = Image.new('RGB', (max_width, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height + image8.height + image9.height + image10.height), color=(255, 255, 255))
    new_image.paste(image1, (center_width_1, 0))
    new_image.paste(image2, (center_width_2, image1.height))
    new_image.paste(image3, (center_width_3, image1.height + image2.height))
    new_image.paste(image4, (center_width_4, image1.height + image2.height + image3.height))
    new_image.paste(image5, (center_width_5, image1.height + image2.height + image3.height + image4.height))
    new_image.paste(image6, (center_width_6, image1.height + image2.height + image3.height + image4.height + image5.height))
    new_image.paste(image7, (center_width_7, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height))
    new_image.paste(image8, (center_width_7, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height))
    new_image.paste(image9, (center_width_7, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height + image8.height))
    new_image.paste(image10, (center_width_7, image1.height + image2.height + image3.height + image4.height + image5.height + image6.height + image7.height + image8.height + image9.height))
    
  if new_image.width > 780:
    diff_width = new_image.width - 780
    diff_ratio = 1 - (diff_width / new_image.width)
  else:
    diff_ratio = 1
    
  img_resize = new_image.resize((int(new_image.width * diff_ratio), int(new_image.height * diff_ratio))) # 이미지 기본 너비 780px 이하로 축소
    
  if cnt < 10:
    new_image.save('img/detail_0' + str(cnt) + '.jpg', 'JPEG', quality=100, subsampling=0)
    img_resize.save('img/detail_0' + str(cnt) + '.jpg', 'JPEG', quality=83, subsampling=0)
  else:
    new_image.save('img/detail_' + str(cnt) + '.jpg', 'JPEG', quality=100, subsampling=0)
    img_resize.save('img/detail_' + str(cnt) + '.jpg', 'JPEG', quality=83, subsampling=0)
    
  # img_resize.save('img/' + str(file_num) + '.jpg', 'JPEG', quality=95, subsampling=0)
  # img_resize.save('img/' + str(file_num) + '_' + file_name + '.jpg', 'JPEG', quality=95, subsampling=0)
  # img_resize.save('img/' + str(code_name) + '_' + file_name + '.jpg', 'JPEG', quality=100, subsampling=0)
  # img_resize.save('img/' + file_name + '.jpg', 'JPEG', quality=100, subsampling=0)
  # img_resize.save('img/' + file_num + '_' + file_name + '.jpg', 'JPEG', quality=90, subsampling=0)
  
  cnt += 1

  for i in range(0, len(img_list)):
    os.remove('./' + str(i) + '.jpg')