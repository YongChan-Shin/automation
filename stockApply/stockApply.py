from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import os

wb = load_workbook('판매데이터.xlsx')
sheet1Name = '판매량 체크'

currPath = os.getcwd()

firstCell = 2
lastCell = wb[sheet1Name].max_row + 1

# 판매정보 생성
sellList = {}
for i in range(firstCell, lastCell):
  sellList[wb[sheet1Name].cell(i, 1).value.replace("/", " ")] = wb[sheet1Name].cell(i, 2).value

print(sellList)

# 전체 재고리스트
prdList = [] 

wb2 = load_workbook('재고 및 품절 데이터.xlsx')

for wb2Sheet in wb2:
  wb2FirstCell = 3
  wb2LastCell = wb2Sheet.max_row + 1
  
  for i in range(wb2FirstCell, wb2LastCell):
    if wb2Sheet.cell(i, 13).value != None:
      prdList.append(wb2Sheet.cell(i, 13).value)    
    if sellList.get(wb2Sheet.cell(i, 13).value) != None:
      if wb2Sheet.cell(i, 14).value > 0:
        if sellList[wb2Sheet.cell(i, 13).value] > wb2Sheet.cell(i, 14).value:
          wb2Sheet.cell(i, 14).value = 0
        else:  
          wb2Sheet.cell(i, 14).value -= sellList[wb2Sheet.cell(i, 13).value]
      else:
        wb2Sheet.cell(i, 14).value = 0
    else:
      continue

# 차감 재고 데이터 매칭 오류 정보 저장
f = open("차감 재고 데이터 매칭 오류.txt", "a+")
for i in sellList:
  if i not in prdList:
    print("{} : {}".format(i, sellList[i]))
    f.write("{} : {}\n".format(i, sellList[i]))
f.close()
    
wb2.save(currPath + '\\재고 및 품절 데이터_판매량 반영.xlsx')