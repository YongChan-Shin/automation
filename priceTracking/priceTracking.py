from openpyxl import Workbook
from openpyxl import load_workbook

import json
import source
import switchPrdInfo

# 키즈꼬모 상품정보
kidscomoPrdList = {}
wb = load_workbook("data.xlsx")
ws = wb.active

for i in range(2, ws.max_row + 1):
  kidscomoPrdList[ws.cell(i, 1).value] = ws.cell(i, 2).value
  
# 삼소니 상품정보
samsonyPrdList = {}
defaultPrice = source.defaultPrice
discountRate = source.discountRate

idx = 2

for i in source.htmlSource:
  prdName = i["optionName1"].split('_')[-1]
  addPrice = int(i["price"])
  salePrice = defaultPrice + addPrice
  
  if prdName not in samsonyPrdList:
    samsonyPrdList[prdName] = int(salePrice * (1 - discountRate))

resultWb = Workbook()
resultWs = resultWb.active

resultWs.cell(1, 1).value = "삼소니 상품명"
resultWs.cell(1, 2).value = "키즈꼬모 상품명"
resultWs.cell(1, 3).value = "삼소니 딜가(A)"
resultWs.cell(1, 4).value = "키즈꼬모 딜가(B)"
resultWs.cell(1, 5).value = "차액(A-B)"
resultWs.cell(1, 6).value = "딜 url"
resultWs.cell(2, 6).value = source.dealUrl

resultWs.column_dimensions["A"].width = 20
resultWs.column_dimensions["B"].width = 20
resultWs.column_dimensions["C"].width = 20
resultWs.column_dimensions["D"].width = 20
resultWs.column_dimensions["E"].width = 20

for key, value in samsonyPrdList.items():
  try:
    resultWs.cell(idx, 1).value = key
    resultWs.cell(idx, 2).value = switchPrdInfo.info[key]
    resultWs.cell(idx, 3).value = value
    resultWs.cell(idx, 4).value = kidscomoPrdList[switchPrdInfo.info[key]]
    resultWs.cell(idx, 5).value = resultWs.cell(idx, 3).value - resultWs.cell(idx, 4).value
    idx += 1
  except:
    print(key, value)

resultWb.save("result.xlsx")


# JSON 파일로 저장
jsonData = {}
jsonData['data'] = []

for key, value in samsonyPrdList.items():
  try:
    jsonData['data'].append({
        "samsonyPrdName": key,
        "kidscomoPrdName": switchPrdInfo.info[key],
        "samsonyPrice": value,
        "kidscomoPrice": kidscomoPrdList[switchPrdInfo.info[key]],
        "priceGap": value - kidscomoPrdList[switchPrdInfo.info[key]]
    })
  except:
    print(key, value)
    
with open("tracking.json", "w", encoding="UTF-8") as outfile:
  json.dump(jsonData, outfile, indent=2, ensure_ascii=False)