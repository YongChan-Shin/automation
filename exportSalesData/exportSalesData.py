from openpyxl import Workbook
from openpyxl import load_workbook

import json

wb = load_workbook("판매데이터.xlsx")
ws = wb["상품사진삽입"]

first_row = 2
last_row = ws.max_row + 1

channelList = [] # 운영채널 리스트

# DB 불러오기
import sqlite3

# 운영채널 정보 생성
con = sqlite3.connect('D:/1.업무/10.기타자료/Development/db/productsData.db')
cur = con.cursor()
cur.execute("SELECT Channel from ProductsData")
data = cur.fetchall()
for i in data:
  if i[0] is not None:
    channelList.append(i[0])
con.close()

print(channelList)

# JSON 파일로 저장
jsonData = {}
jsonData["channelList"] = channelList
jsonData["data"] = []

# 상품 호수별 판매정보 생성
salesPrdSizeQnt = {}
for i in range(first_row, last_row):
  if ws.cell(i, 10).value not in salesPrdSizeQnt:
    salesPrdSizeQnt[ws.cell(i, 10).value] = {'{}'.format(ws.cell(i, 11).value): ws.cell(i, 12).value}
  else:
    if ws.cell(i, 11).value not in salesPrdSizeQnt[ws.cell(i, 10).value]:
      salesPrdSizeQnt[ws.cell(i, 10).value][ws.cell(i, 11).value] = ws.cell(i, 12).value
    else:
      salesPrdSizeQnt[ws.cell(i, 10).value][ws.cell(i, 11).value] += ws.cell(i, 12).value
      
print(salesPrdSizeQnt)
  
for i in range(first_row, last_row):
  if ws.cell(row=i, column=1).value == None or ws.cell(row=i, column=1).value == '':
    continue
  else:
    jsonData["data"].append({
      "prdName": ws.cell(i, 1).value.replace("토밍이세트", "토밍이모자세트").replace("해피스노우세트", "해피스노우모자세트"),
      "salesCnt": ws.cell(i, 2).value,
      "sizeQnt": salesPrdSizeQnt[ws.cell(i, 1).value.replace("토밍이세트", "토밍이모자세트").replace("해피스노우세트", "해피스노우모자세트")]
    })
  
print(jsonData)

with open("salesData.json", "w", encoding="UTF-8") as outfile:
  json.dump(jsonData, outfile, indent=2, ensure_ascii=False)