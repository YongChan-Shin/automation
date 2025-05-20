from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

driver.get('https://www.naver.com/')

# 네이버 로그인 후 작업 진행
url = input('작업 시작(대상 url 입력)')
driver.get(url)
driver.switch_to.frame('mainFrame')

imgEls = driver.find_elements(By.CLASS_NAME, 'se-image-resource')

wb = Workbook()
ws = wb.active
ws.cell(1, 1).value = '순번'
ws.cell(1, 2).value = '이미지 url'
ws.column_dimensions['B'].width = 200

for idx, el in enumerate(imgEls):
  ws.cell(idx + 2, 1).value = idx + 1
  ws.cell(idx + 2, 2).value = el.get_attribute('src')
  
wb.save('이미지 url 추출 데이터.xlsx')

time.sleep(100000000)