from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import os

wb = load_workbook('data.xlsx')
sheet1Name = '상품리스트'
sheet2Name = '추가완료'

currPath = os.getcwd()

firstCell = 2
lastCell = wb[sheet1Name].max_row + 1
wb.active = wb[sheet2Name]

wb2 = load_workbook('데이터.xlsx')

prdIdxInfoList = [] # 품절상품 정보
prdIdxInfoCSList = [] # 품절상품(CS팀전달) 정보
stockList = {} # 상품별 재고수량 정보
channelList = {} # 상품별 세팅채널 정보

firstCellInfoCS = 3
lastCellInfoCs = wb2['품절상품(CS팀전달)'].max_row + 1

for i in range(firstCellInfoCS, lastCellInfoCs):
  prdIdxInfoCSList.append(wb2['품절상품(CS팀전달)'].cell(i, 17).value)

for wb2Sheet in wb2:
  wb2FirstCell = 3
  wb2LastCell = wb2Sheet.max_row + 1

  # 품절상품 정보 리스트 생성
  for row in wb2Sheet.iter_rows(min_row=wb2FirstCell, max_row=wb2LastCell, min_col=17, max_col=17):
      for cell in row:
          prdIdxInfoList.append(cell.value)
          
  fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
  fillData2 = PatternFill(fill_type='solid', start_color='DDEBF7', end_color='DDEBF7')
  fillFont = Font(color='FF0000')

  for stockCnt in range(wb2FirstCell, wb2LastCell):
    if wb2Sheet.cell(stockCnt, 13).value == None or wb2Sheet.cell(stockCnt, 13).value == "":
      continue
    # 상품식별값에 따른 재고수량 매칭
    stockList[wb2Sheet.cell(stockCnt, 13).value.replace(" ", "")] = wb2Sheet.cell(stockCnt, 14).value
    channelList[wb2Sheet.cell(stockCnt, 13).value.replace(" ", "")] = wb2Sheet.cell(stockCnt, 19).value

  cnt = 2

  for i in range(firstCell, lastCell):
    if wb[sheet1Name].cell(i, 4).value == None or wb[sheet1Name].cell(i, 4).value == '':
      continue
    colorList = wb[sheet1Name].cell(i, 4).value.split('/')
    for color in colorList:
      sizeList = wb[sheet1Name].cell(i, 5).value.split('/')
      for size in sizeList:
        for j in range(1, 16):
          wb[sheet2Name].cell(cnt, j).value = wb[sheet1Name].cell(i, j).value
        wb[sheet2Name].cell(cnt, 4).value = color
        wb[sheet2Name].cell(cnt, 5).value = size
        wb[sheet2Name].cell(cnt, 6).value = wb[sheet2Name].cell(cnt, 2).value + " " + color + " " + size
        prdIdxInfo = wb[sheet2Name].cell(cnt, 6).value
        try:
          wb[sheet2Name].cell(cnt, 7).value = stockList[prdIdxInfo.replace(" ", "").replace("(저스틴23)", "").replace("(주니어)", "")]
          wb[sheet2Name].cell(cnt, 9).value = channelList[prdIdxInfo.replace(" ", "").replace("(저스틴23)", "").replace("(주니어)", "")]
          if wb[sheet2Name].cell(cnt, 9).value == "랜덤박스":
            wb[sheet2Name].cell(cnt, 7).value = 0
          wb[sheet2Name].cell(cnt, 10).value = " "
        except:
          pass
        
        try:
          if wb[sheet2Name].cell(cnt, 7).value >= 10000:
            wb[sheet2Name].cell(cnt, 7).value = 50
        except:
          pass
        
        if prdIdxInfo in prdIdxInfoList or wb[sheet2Name].cell(cnt, 7).value == 0:
          if prdIdxInfo in prdIdxInfoCSList:
            for z in range(1, 16):
              wb[sheet2Name].cell(cnt, z).fill = fillData
              wb[sheet2Name].cell(cnt, z).font = fillFont
            wb[sheet2Name].cell(cnt, 8).value = "품절"
          else:
            for z in range(1, 16):
              wb[sheet2Name].cell(cnt, z).fill = fillData2
              wb[sheet2Name].cell(cnt, z).font = fillFont
            wb[sheet2Name].cell(cnt, 8).value = "자동품절(판매량차감)"
            
        
        cnt += 1
  wb[sheet2Name].auto_filter.ref = "A1:O1"
  
for sheet in wb:
  if sheet.title == sheet2Name:
    sheet.sheet_view.tabSelected = True
  else:
    sheet.sheet_view.tabSelected = False

wb.save(currPath + '\\data_추가완료.xlsx')