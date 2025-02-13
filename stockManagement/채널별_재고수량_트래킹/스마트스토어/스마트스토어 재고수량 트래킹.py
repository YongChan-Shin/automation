from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
# import os
# from os import listdir
# from os.path import exists
# from os import makedirs

# 재고정보 생성
wbStock = load_workbook('데이터.xlsx')

stockList = {} # 재고정보
stockErrList = [] # 품절상품 중 판매세팅된 상품정보
stockErrAutoList = [] # 품절상품 중 판매세팅된 상품정보(자동품절건)

soldoutPrdCSList = [] # 품절상품(CS팀전달)

impendingPrdList = [] # 재고 보충 필요 상품정보

matchingErrList = [] # 상품정보 매칭 오류건

first_row_cs = 3
last_row_cs = wbStock['품절상품(CS팀전달)'].max_row + 1

for i in range(first_row_cs, last_row_cs):
  soldoutPrdCSList.append(wbStock['품절상품(CS팀전달)'].cell(i, 17).value)

for wbSheet in wbStock:
  wbFirstCell = 3
  wbLastCell = wbSheet.max_row + 1
  
  for i in range(wbFirstCell, wbLastCell):
    if wbSheet.cell(i, 13).value != None:
      stockList[wbSheet.cell(i, 13).value] = wbSheet.cell(i, 14).value

# 상품정보 리스트
product_list = []
color_list = []
size_list = []
cap_list = []
excProducts = [] # 판매중지상품 판매여부 체크용

# DB 불러오기
import sqlite3
con = sqlite3.connect('D:/1.업무/10.기타자료/Development/db/productsData.db')
cur = con.cursor()

cur.execute("SELECT PrdName from ProductsData WHERE PrdName IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  product_list.append(i[0])

cur.execute("SELECT Color from ProductsData WHERE Color IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  color_list.append(i[0])

cur.execute("SELECT Size from ProductsData WHERE Size IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  size_list.append(i[0])
  
cur.execute("SELECT Cap from ProductsData WHERE Cap IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  cap_list.append(i[0])

cur.execute("SELECT ExcProducts from ProductsData WHERE ExcProducts IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  excProducts.append(i[0])
  
# 판매중지상품 판매여부 체크용
excProductsCheckList = []

wb = load_workbook('./옵션.xlsx')

wb.create_sheet('정리')
ws_name = wb.get_sheet_names()

sheet1 = wb[str(ws_name[0])]
sheet2 = wb[str(ws_name[1])]

first_row_sh1 = 6
first_row_sh2 = 2
last_row_sh1 = sheet1.max_row + 1

fillDataHeader = PatternFill(fill_type='solid', start_color='E7E7E7', end_color='E7E7E7')
fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillData2 = PatternFill(fill_type='solid', start_color='FFBDBD', end_color='FFBDBD')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True)

sheet2.cell(1, 1).value = "상품번호"
sheet2.cell(1, 2).value = "상품정보"
sheet2.cell(1, 3).value = "상품명"
sheet2.cell(1, 4).value = "컬러"
sheet2.cell(1, 5).value = "사이즈"
sheet2.cell(1, 6).value = "주문정보 정제"
sheet2.cell(1, 7).value = "옵션사용여부"
sheet2.cell(1, 8).value = "옵션재고수량"
sheet2.cell(1, 9).value = "재고(데이터파일 기준)"

sheet2.cell(1, 1).alignment = fillAlignment
sheet2.cell(1, 2).alignment = fillAlignment
sheet2.cell(1, 3).alignment = fillAlignment
sheet2.cell(1, 4).alignment = fillAlignment
sheet2.cell(1, 5).alignment = fillAlignment
sheet2.cell(1, 6).alignment = fillAlignment
sheet2.cell(1, 7).alignment = fillAlignment
sheet2.cell(1, 8).alignment = fillAlignment
sheet2.cell(1, 9).alignment = fillAlignment

sheet2.cell(1, 1).font = fillFont
sheet2.cell(1, 2).font = fillFont
sheet2.cell(1, 3).font = fillFont
sheet2.cell(1, 4).font = fillFont
sheet2.cell(1, 5).font = fillFont
sheet2.cell(1, 6).font = fillFont
sheet2.cell(1, 7).font = fillFont
sheet2.cell(1, 8).font = fillFont
sheet2.cell(1, 9).font = fillFont

sheet2.cell(1, 1).fill = fillDataHeader
sheet2.cell(1, 2).fill = fillDataHeader
sheet2.cell(1, 3).fill = fillData
sheet2.cell(1, 4).fill = fillData
sheet2.cell(1, 5).fill = fillData
sheet2.cell(1, 6).fill = fillData
sheet2.cell(1, 7).fill = fillData
sheet2.cell(1, 8).fill = fillData
sheet2.cell(1, 9).fill = fillData

sheet2.column_dimensions['A'].width = 20
sheet2.column_dimensions['B'].width = 40
sheet2.column_dimensions['C'].width = 30
sheet2.column_dimensions['D'].width = 20
sheet2.column_dimensions['E'].width = 20
sheet2.column_dimensions['F'].width = 40
sheet2.column_dimensions['G'].width = 20
sheet2.column_dimensions['H'].width = 20
sheet2.column_dimensions['I'].width = 20

for i in range(first_row_sh1, last_row_sh1):
  try:
    optList = sheet1.cell(i, 11).value.split('\n')
    optStock = sheet1.cell(i, 13).value.split('\n')
    optIsUse = sheet1.cell(i, 14).value.split('\n')
    idx = 0
    for opt in optList:
      last_row_sh2 = sheet2.max_row + 1
      sheet2.cell(last_row_sh2, 1).value = sheet1.cell(i, 1).value
      sheet2.cell(last_row_sh2, 2).value = sheet1.cell(i, 10).value + "/" + sheet1.cell(i, 11).value.split('\n')[idx]
      sheet2.cell(last_row_sh2, 7).value = optIsUse[idx]
      sheet2.cell(last_row_sh2, 8).value = optStock[idx]
      idx += 1
  except:
    pass

for i in range(first_row_sh2, last_row_sh2 + 1):
  try:
    
    for product in product_list:
      try:
        if product in str(sheet2.cell(row=i, column=2).value):
          prdDetailInfoProduct = product.replace("(저스틴23)", "")
          sheet2.cell(row=i, column=3).value = prdDetailInfoProduct
      except:
        pass
    for color in color_list:
      if color in str(sheet2.cell(row=i, column=2).value):
        prdDetailInfoColor = color
        sheet2.cell(row=i, column=4).value = prdDetailInfoColor
    for size in size_list:
      if size in str(sheet2.cell(row=i, column=2).value):
        prdDetailInfoSize = size.replace("FREE", "free")
        sheet2.cell(row=i, column=5).value = prdDetailInfoSize
        
    if prdDetailInfoProduct in cap_list:
        prdDetailInfoSize = "free"
    
    prdDetailInfo = '{} {} {}'.format(prdDetailInfoProduct, prdDetailInfoColor, prdDetailInfoSize)
    
    sheet2.cell(i, 6).value = prdDetailInfo
    sheet2.cell(i, 9).value = stockList[sheet2.cell(i, 6).value]
    
    if stockList[sheet2.cell(i, 6).value] == 0:
      if sheet2.cell(row=i, column=7).value == "Y":
        if int(sheet2.cell(row=i, column=8).value) != 0:
          print("{}/{}".format(sheet2.cell(i, 6).value, sheet2.cell(i, 9).value))
          if sheet2.cell(i, 6).value in soldoutPrdCSList:
            stockErrList.append("○ 상품번호 : {} / {} / 옵션사용여부 : {} / 옵션재고수량 : {} / 데이터파일 기준 재고 : 0".format(sheet2.cell(i, 1).value, sheet2.cell(i, 6).value, sheet2.cell(i, 7).value, sheet2.cell(i, 8).value))
          else:
            stockErrAutoList.append("※ 판매량차감 자동품절 상품(CS팀에서 품절로 전달되지 않은 상품) ※\n○ 상품번호 : {} / {} / 옵션사용여부 : {} / 옵션재고수량 : {} / 데이터파일 기준 재고 : 0".format(sheet2.cell(i, 1).value, sheet2.cell(i, 6).value, sheet2.cell(i, 7).value, sheet2.cell(i, 8).value))         
          for colNum in range(1, 10):
            sheet2.cell(row=i, column=colNum).fill = fillData2
            
    if stockList[sheet2.cell(i, 6).value] != 0:
      if sheet2.cell(row=i, column=7).value == "Y":
        if int(sheet2.cell(row=i, column=8).value) <= 3:
          if stockList[sheet2.cell(i, 6).value] > int(sheet2.cell(row=i, column=8).value):
            impendingPrdList.append("○ 상품번호 : {} / {} / 옵션사용여부 : {} / 옵션재고수량 : {} / 데이터파일 기준 재고 : {}".format(sheet2.cell(i, 1).value, sheet2.cell(i, 6).value, sheet2.cell(i, 7).value, sheet2.cell(i, 8).value, stockList[sheet2.cell(i, 6).value]))
          
      if sheet2.cell(row=i, column=7).value == "Y":
        if int(sheet2.cell(row=i, column=8).value) != 0:
          if prdDetailInfoProduct in excProducts:
            excProductsCheckList.append("○ 상품번호 : {} / {} / 옵션사용여부 : {} / 옵션재고수량 : {}".format(sheet2.cell(i, 1).value, sheet2.cell(i, 6).value, sheet2.cell(i, 7).value, sheet2.cell(i, 8).value))
            
  except Exception as e:
    matchingErrList.append('{} / {}'.format(prdDetailInfo, e))
    continue
  
  
if len(stockErrList) > 0 or len(stockErrAutoList) > 0:
  f = open("(스마트스토어) 품절상품 중 판매세팅된 상품 정보.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(스마트스토어) 품절상품 중 판매세팅된 상품 정보\n\n")
  for i in stockErrList:
    f.write("{}\n\n".format(i))
  for i in stockErrAutoList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(impendingPrdList) > 0:
  f = open("(스마트스토어) 재고 보충 필요 상품 정보(품절 혹은 품절임박).txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(스마트스토어) 재고 보충 필요 상품 정보(품절 혹은 품절임박)\n\n")
  for i in impendingPrdList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(excProductsCheckList) > 0:
  f = open("(스마트스토어) 판매제외 상품 포함 체크.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(스마트스토어) 판매제외 상품 포함 체크\n\n")
  for i in excProductsCheckList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(matchingErrList) > 0:
  f = open("(스마트스토어) 상품정보 매칭 오류건.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(스마트스토어) 상품정보 매칭 오류건\n\n")
  for i in matchingErrList:
    f.write("{}\n\n".format(i))
  f.close()  

sheet2.auto_filter.ref = "A1:I1"


for sheet in wb:
  if sheet.title == '정리':
    sheet.sheet_view.tabSelected = True
  else:
    sheet.sheet_view.tabSelected = False
    
wb.active = sheet2

wb.save('상품옵션별 재고현황 추출.xlsx')