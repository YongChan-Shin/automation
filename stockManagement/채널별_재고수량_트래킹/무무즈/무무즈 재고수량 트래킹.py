from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import os
import json
import datetime
# from os import listdir
# from os.path import exists
# from os import makedirs

# 판매세팅 상품정보 JSON 파일 불러오기
channelName = '무무즈'
# settingInfoJSONPath = os.path.dirname(os.path.abspath(os.path.dirname(__file__))) + '\\settingProducts.json'
settingInfoJSONPath = 'D:\\1.업무\\6.자동화작업\\재고 관리\\채널별_재고수량_트래킹\\settingProducts.json'
with open(settingInfoJSONPath, 'r', encoding='UTF-8') as jsonFile:
  settingInfo = json.load(jsonFile)

# 재고정보 생성
wbStock = load_workbook('데이터.xlsx')

stockList = {} # 재고정보
stockErrList = [] # 품절상품 중 판매세팅된 상품정보
stockErrAutoList = [] # 품절상품 중 판매세팅된 상품정보(자동품절건)

soldoutPrdCSList = [] # 품절상품(CS팀전달)

impendingPrdList = [] # 재고 보충 필요 상품정보

matchingErrList = [] # 상품정보 매칭 오류건

first_row_cs = 3
last_row_cs = wbStock['품절상품(CS팀전달)'].max_row + 1

for i in range(first_row_cs, last_row_cs):
  soldoutPrdCSList.append(wbStock['품절상품(CS팀전달)'].cell(i, 17).value)

for wbSheet in wbStock:
  wbFirstCell = 3
  wbLastCell = wbSheet.max_row + 1
  
  for i in range(wbFirstCell, wbLastCell):
    if wbSheet.cell(i, 13).value != None:
      stockList[wbSheet.cell(i, 13).value] = wbSheet.cell(i, 14).value

# 상품정보 리스트
product_list = []
color_list = []
size_list = []
cap_list = []
excProducts = [] # 판매중지상품 판매여부 체크용

# DB 불러오기
import sqlite3
con = sqlite3.connect('D:/1.업무/10.기타자료/Development/db/productsData.db')
cur = con.cursor()

cur.execute("SELECT PrdName from ProductsData WHERE PrdName IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  product_list.append(i[0])

cur.execute("SELECT Color from ProductsData WHERE Color IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  color_list.append(i[0])

cur.execute("SELECT Size from ProductsData WHERE Size IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  size_list.append(i[0])
  
cur.execute("SELECT Cap from ProductsData WHERE Cap IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  cap_list.append(i[0])

cur.execute("SELECT ExcProducts from ProductsData WHERE ExcProducts IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  excProducts.append(i[0])
  
# 판매중지상품 판매여부 체크용
excProductsCheckList = []

wb = load_workbook('./옵션.xlsx')

wb.create_sheet('정리')
ws_name = wb.get_sheet_names()

sheet1 = wb[str(ws_name[0])]
sheet2 = wb[str(ws_name[1])]

first_row = 2
last_row_sh1 = sheet1.max_row + 1

fillDataHeader = PatternFill(fill_type='solid', start_color='E7E7E7', end_color='E7E7E7')
fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillData2 = PatternFill(fill_type='solid', start_color='FFBDBD', end_color='FFBDBD')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True)

sheet2.cell(1, 1).value = "상품명"
sheet2.cell(1, 2).value = "컬러"
sheet2.cell(1, 3).value = "사이즈"
sheet2.cell(1, 4).value = "노출상태"
sheet2.cell(1, 5).value = "옵션노출상태"
sheet2.cell(1, 6).value = "옵션판매상태"
sheet2.cell(1, 7).value = "옵션품절상태"
sheet2.cell(1, 8).value = '상품정보'
sheet2.cell(1, 9).value = '상품명'
sheet2.cell(1, 10).value = '컬러'
sheet2.cell(1, 11).value = '사이즈'
sheet2.cell(1, 12).value = '주문정보 정제'
sheet2.cell(1, 13).value = '재고(데이터파일 기준)'

sheet2.cell(1, 1).alignment = fillAlignment
sheet2.cell(1, 2).alignment = fillAlignment
sheet2.cell(1, 3).alignment = fillAlignment
sheet2.cell(1, 4).alignment = fillAlignment
sheet2.cell(1, 5).alignment = fillAlignment
sheet2.cell(1, 6).alignment = fillAlignment
sheet2.cell(1, 7).alignment = fillAlignment
sheet2.cell(1, 8).alignment = fillAlignment
sheet2.cell(1, 9).alignment = fillAlignment
sheet2.cell(1, 10).alignment = fillAlignment
sheet2.cell(1, 11).alignment = fillAlignment
sheet2.cell(1, 12).alignment = fillAlignment
sheet2.cell(1, 13).alignment = fillAlignment

sheet2.cell(1, 1).font = fillFont
sheet2.cell(1, 2).font = fillFont
sheet2.cell(1, 3).font = fillFont
sheet2.cell(1, 4).font = fillFont
sheet2.cell(1, 5).font = fillFont
sheet2.cell(1, 6).font = fillFont
sheet2.cell(1, 7).font = fillFont
sheet2.cell(1, 8).font = fillFont
sheet2.cell(1, 9).font = fillFont
sheet2.cell(1, 10).font = fillFont
sheet2.cell(1, 11).font = fillFont
sheet2.cell(1, 12).font = fillFont
sheet2.cell(1, 13).font = fillFont

sheet2.cell(1, 1).fill = fillDataHeader
sheet2.cell(1, 2).fill = fillDataHeader
sheet2.cell(1, 3).fill = fillDataHeader
sheet2.cell(1, 4).fill = fillDataHeader
sheet2.cell(1, 5).fill = fillDataHeader
sheet2.cell(1, 6).fill = fillDataHeader
sheet2.cell(1, 7).fill = fillDataHeader
sheet2.cell(1, 8).fill = fillData
sheet2.cell(1, 9).fill = fillData
sheet2.cell(1, 10).fill = fillData
sheet2.cell(1, 11).fill = fillData
sheet2.cell(1, 12).fill = fillData
sheet2.cell(1, 13).fill = fillData

sheet2.column_dimensions['A'].width = 20
sheet2.column_dimensions['B'].width = 20
sheet2.column_dimensions['C'].width = 20
sheet2.column_dimensions['D'].width = 20
sheet2.column_dimensions['E'].width = 20
sheet2.column_dimensions['F'].width = 20
sheet2.column_dimensions['G'].width = 20

sheet2.column_dimensions['H'].width = 40
sheet2.column_dimensions['I'].width = 25
sheet2.column_dimensions['J'].width = 25
sheet2.column_dimensions['K'].width = 25
sheet2.column_dimensions['L'].width = 40
sheet2.column_dimensions['M'].width = 25

for i in range(first_row, last_row_sh1):
  try:
    optList = sheet1.cell(i, 56).value.split('\n')
    optSoldoutList = sheet1.cell(i, 83).value.split('\n')
    idx = 0
    for opt in optList:
      last_row_sh2 = sheet2.max_row + 1
      sheet2.cell(last_row_sh2, 1).value = sheet1.cell(i, 13).value
      sheet2.cell(last_row_sh2, 2).value = opt.split('/')[0]
      sheet2.cell(last_row_sh2, 3).value = opt.split('/')[1]
      sheet2.cell(last_row_sh2, 4).value = sheet1.cell(i, 8).value
      sheet2.cell(last_row_sh2, 5).value = opt.split('/')[-3]
      sheet2.cell(last_row_sh2, 6).value = opt.split('/')[-2]
      sheet2.cell(last_row_sh2, 7).value = optSoldoutList[idx]
      idx += 1
  except:
    pass

for i in range(first_row, last_row_sh2 + 1):
  try:
    sheet2.cell(i, 8).value = str(sheet2.cell(i, 1).value) + '/' + str(sheet2.cell(i, 2).value) + '/' + str(sheet2.cell(i, 3).value)
    sheet2.cell(i, 8).value = sheet2.cell(i, 8).value.replace(" ", "")
    
    for product in product_list:
      try:
        if product in str(sheet2.cell(row=i, column=8).value):
          prdDetailInfoProduct = product.replace("(저스틴23)", "")
          sheet2.cell(row=i, column=9).value = prdDetailInfoProduct
      except:
        pass
    for color in color_list:
      if color in str(sheet2.cell(row=i, column=8).value):
        prdDetailInfoColor = color
        sheet2.cell(row=i, column=10).value = prdDetailInfoColor
    for size in size_list:
      if size in str(sheet2.cell(row=i, column=8).value):
        prdDetailInfoSize = size.replace("FREE", "free")
        sheet2.cell(row=i, column=11).value = prdDetailInfoSize
        
    if prdDetailInfoProduct in cap_list:
        prdDetailInfoSize = "free"
    
    prdDetailInfo = '{} {} {}'.format(prdDetailInfoProduct, prdDetailInfoColor, prdDetailInfoSize)
    
    sheet2.cell(i, 12).value = prdDetailInfo
    sheet2.cell(i, 13).value = stockList[sheet2.cell(i, 12).value]
    
    if stockList[sheet2.cell(i, 12).value] == 0:
      if sheet2.cell(row=i, column=4).value == "노출함":
        if sheet2.cell(row=i, column=5).value == "노출함" and sheet2.cell(row=i, column=6).value == "판매함":
          if sheet2.cell(row=i, column=7).value != "품절":
            print("{}/{}".format(sheet2.cell(i, 8).value, sheet2.cell(i, 7).value))
            if sheet2.cell(i, 8).value in soldoutPrdCSList:
              stockErrList.append("○ {} / 노출상태 : {} / 옵션노출상태 : {} / 옵션판매상태 : {} / 옵션품절상태 : {} / 데이터파일 기준 재고 : 0".format(sheet2.cell(i, 12).value, sheet2.cell(i, 4).value, sheet2.cell(i, 5).value, sheet2.cell(i, 6).value, sheet2.cell(i, 7).value))
            else:
              stockErrAutoList.append("※ 판매량차감 자동품절 상품(CS팀에서 품절로 전달되지 않은 상품) ※\n○ {} / 노출상태 : {} / 옵션노출상태 : {} / 옵션판매상태 : {} / 옵션품절상태 : {} / 데이터파일 기준 재고 : 0".format(sheet2.cell(i, 12).value, sheet2.cell(i, 4).value, sheet2.cell(i, 5).value, sheet2.cell(i, 6).value, sheet2.cell(i, 7).value))              
            for colNum in range(1, 14):
              sheet2.cell(row=i, column=colNum).fill = fillData2
              
    if stockList[sheet2.cell(i, 12).value] != 0:
      if sheet2.cell(row=i, column=7).value == "품절" or sheet2.cell(row=i, column=7).value == "임시품절":
        if sheet2.cell(row=i, column=6).value == "판매함":
          impendingPrdList.append("○ {} / 노출상태 : {} / 옵션노출상태 : {} / 옵션판매상태 : {} / 옵션품절상태 : {} / 데이터파일 기준 재고 : {}".format(sheet2.cell(i, 12).value, sheet2.cell(i, 4).value, sheet2.cell(i, 5).value, sheet2.cell(i, 6).value, sheet2.cell(i, 7).value, stockList[sheet2.cell(i, 12).value]))
          
    if sheet2.cell(row=i, column=4).value == "노출함":
      if sheet2.cell(row=i, column=5).value == "노출함" and sheet2.cell(row=i, column=6).value == "판매함":
        if sheet2.cell(row=i, column=7).value != "품절":
          
          # 판매세팅 상품정보 추가
          settingInfo['checkTime'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
          
          if sheet2.cell(i, 9).value not in settingInfo:
            settingInfo[sheet2.cell(i, 9).value] = [channelName]
          else:
            if channelName not in settingInfo[sheet2.cell(i, 9).value]:
              settingInfo[sheet2.cell(i, 9).value].append(channelName)
              
          if prdDetailInfoProduct in excProducts:
            excProductsCheckList.append("○ {} / 노출상태 : {} / 옵션노출상태 : {} / 옵션판매상태 : {} / 옵션품절상태 : {}".format(sheet2.cell(i, 12).value, sheet2.cell(i, 4).value, sheet2.cell(i, 5).value, sheet2.cell(i, 6).value, sheet2.cell(i, 7).value))
  except Exception as e:
    matchingErrList.append('row : {} / {}'.format(i, e))
    continue
  
  
if len(stockErrList) > 0 or len(stockErrAutoList) > 0:
  f = open("(무무즈) 품절상품 중 판매세팅된 상품 정보.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(무무즈) 품절상품 중 판매세팅된 상품 정보\n\n")
  for i in stockErrList:
    f.write("{}\n\n".format(i))
  for i in stockErrAutoList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(impendingPrdList) > 0:
  f = open("(무무즈) 재고 보충 필요 상품 정보(품절 혹은 품절임박).txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(무무즈) 재고 보충 필요 상품 정보(품절 혹은 품절임박)\n\n")
  for i in impendingPrdList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(excProductsCheckList) > 0:
  f = open("(무무즈) 판매제외 상품 포함 체크.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(무무즈) 판매제외 상품 포함 체크\n\n")
  for i in excProductsCheckList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(matchingErrList) > 0:
  f = open("(무무즈) 상품정보 매칭 오류건.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(무무즈) 상품정보 매칭 오류건\n\n")
  for i in matchingErrList:
    f.write("{}\n\n".format(i))
  f.close()  

sheet2.auto_filter.ref = "A1:M1"

# 판매세팅 상품정보 JSON 파일 업데이트
with open(settingInfoJSONPath, 'w', encoding='UTF-8') as jsonFile:
  json.dump(settingInfo, jsonFile, indent=2, ensure_ascii=False)

for sheet in wb:
  if sheet.title == '정리':
    sheet.sheet_view.tabSelected = True
  else:
    sheet.sheet_view.tabSelected = False
    
wb.active = sheet2

wb.save('상품옵션별 재고현황 추출.xlsx')