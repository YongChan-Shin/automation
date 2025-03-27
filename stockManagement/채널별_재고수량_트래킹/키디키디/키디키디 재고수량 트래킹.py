from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import os
import json
# import os
# from os import listdir
# from os.path import exists
# from os import makedirs

# 판매세팅 상품정보 JSON 파일 불러오기
channelName = '키디키디'
# settingInfoJSONPath = os.path.dirname(os.path.abspath(os.path.dirname(__file__))) + '\\settingProducts.json'
settingInfoJSONPath = 'D:\\1.업무\\6.자동화작업\\재고 관리\\채널별_재고수량_트래킹\\settingProducts.json'
with open(settingInfoJSONPath, 'r', encoding='UTF-8') as jsonFile:
  settingInfo = json.load(jsonFile)

# 재고정보 생성
wbStock = load_workbook('데이터.xlsx')

stockList = {} # 재고정보
stockErrList = [] # 품절상품 중 판매세팅된 상품정보
stockErrAutoList = [] # 품절상품 중 판매세팅된 상품정보(자동품절건)

soldoutPrdCSList = [] # 품절상품(CS팀전달)

impendingPrdList = [] # 재고 보충 필요 상품정보

matchingErrList = [] # 상품정보 매칭 오류건

first_row_cs = 3
last_row_cs = wbStock['품절상품(CS팀전달)'].max_row + 1

for i in range(first_row_cs, last_row_cs):
  soldoutPrdCSList.append(wbStock['품절상품(CS팀전달)'].cell(i, 17).value)

for wbSheet in wbStock:
  wbFirstCell = 3
  wbLastCell = wbSheet.max_row + 1
  
  for i in range(wbFirstCell, wbLastCell):
    if wbSheet.cell(i, 13).value != None:
      stockList[wbSheet.cell(i, 13).value] = wbSheet.cell(i, 14).value

# 상품정보 리스트
product_list = []
color_list = []
size_list = []
cap_list = []
excProducts = [] # 판매중지상품 판매여부 체크용

# DB 불러오기
import sqlite3
con = sqlite3.connect('D:/1.업무/10.기타자료/Development/db/productsData.db')
cur = con.cursor()

cur.execute("SELECT PrdName from ProductsData WHERE PrdName IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  product_list.append(i[0])

cur.execute("SELECT Color from ProductsData WHERE Color IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  color_list.append(i[0])

cur.execute("SELECT Size from ProductsData WHERE Size IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  size_list.append(i[0])
  
cur.execute("SELECT Cap from ProductsData WHERE Cap IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  cap_list.append(i[0])

cur.execute("SELECT ExcProducts from ProductsData WHERE ExcProducts IS NOT NULL ORDER BY rowid")
data = cur.fetchall()
for i in data:
  excProducts.append(i[0])

print(excProducts)
# 판매중지상품 판매여부 체크용
excProductsCheckList = []

wb = load_workbook('./옵션.xlsx')
ws = wb.active

fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillData2 = PatternFill(fill_type='solid', start_color='FFBDBD', end_color='FFBDBD')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True)

ws.cell(1, 13).value = '상품정보'
ws.cell(1, 14).value = '상품명'
ws.cell(1, 15).value = '컬러'
ws.cell(1, 16).value = '사이즈'
ws.cell(1, 17).value = '주문정보 정제'
ws.cell(1, 18).value = '재고(데이터파일 기준)'

ws.cell(1, 13).alignment = fillAlignment
ws.cell(1, 14).alignment = fillAlignment
ws.cell(1, 15).alignment = fillAlignment
ws.cell(1, 16).alignment = fillAlignment
ws.cell(1, 17).alignment = fillAlignment
ws.cell(1, 18).alignment = fillAlignment

ws.cell(1, 13).font = fillFont
ws.cell(1, 14).font = fillFont
ws.cell(1, 15).font = fillFont
ws.cell(1, 16).font = fillFont
ws.cell(1, 17).font = fillFont
ws.cell(1, 18).font = fillFont

ws.cell(1, 13).fill = fillData
ws.cell(1, 14).fill = fillData
ws.cell(1, 15).fill = fillData
ws.cell(1, 16).fill = fillData
ws.cell(1, 17).fill = fillData
ws.cell(1, 18).fill = fillData

ws.column_dimensions['M'].width = 40
ws.column_dimensions['N'].width = 25
ws.column_dimensions['O'].width = 25
ws.column_dimensions['P'].width = 25
ws.column_dimensions['Q'].width = 40
ws.column_dimensions['R'].width = 25


first_row = 2
last_row = ws.max_row + 1

for i in range(first_row, last_row):
  try:
    ws.cell(i, 13).value = str(ws.cell(i, 3).value) + ' / ' + str(ws.cell(i, 5).value)
    ws.cell(i, 13).value = ws.cell(i, 13).value.replace(" ", "")
    
    for product in product_list:
      try:
        if product in str(ws.cell(row=i, column=13).value):
          prdDetailInfoProduct = product.replace("(저스틴23)", "")
          ws.cell(row=i, column=14).value = prdDetailInfoProduct
      except:
        pass
    for color in color_list:
      if color in str(ws.cell(row=i, column=13).value):
        prdDetailInfoColor = color
        ws.cell(row=i, column=15).value = prdDetailInfoColor
    for size in size_list:
      if size in str(ws.cell(row=i, column=13).value):
        prdDetailInfoSize = size.replace("FREE", "free")
        ws.cell(row=i, column=16).value = prdDetailInfoSize
        
    if prdDetailInfoProduct in cap_list:
        prdDetailInfoSize = "free"
    
    prdDetailInfo = '{} {} {}'.format(prdDetailInfoProduct, prdDetailInfoColor, prdDetailInfoSize)
    
    ws.cell(i, 17).value = prdDetailInfo
    ws.cell(i, 18).value = stockList[ws.cell(i, 17).value]
    
    if stockList[ws.cell(i, 17).value] == 0:
      if ws.cell(row=i, column=6).value == "판매진행":
        if int(ws.cell(row=i, column=8).value) != 0:
          print("{}/{}".format(ws.cell(i, 17).value, stockList[ws.cell(i, 17).value]))
          if ws.cell(i, 17).value in soldoutPrdCSList:
            stockErrList.append("○ {} / 상품번호 : {} / 판매상태 : {} / 재고수량 : {} / 데이터파일 기준 재고 : 0".format(ws.cell(i, 17).value, ws.cell(i, 2).value, ws.cell(i, 6).value, ws.cell(i, 8).value))
          else:
            stockErrAutoList.append("※ 판매량차감 자동품절 상품(CS팀에서 품절로 전달되지 않은 상품) ※\n○ {} / 상품번호 : {} / 판매상태 : {} / 재고수량 : {} / 데이터파일 기준 재고 : 0".format(ws.cell(i, 17).value, ws.cell(i, 2).value, ws.cell(i, 6).value, ws.cell(i, 8).value))
          for colNum in range(1, 19):
            ws.cell(row=i, column=colNum).fill = fillData2
            
    if stockList[ws.cell(i, 17).value] != 0:
      if ws.cell(row=i, column=6).value == "판매진행":
        if int(ws.cell(row=i, column=8).value) <= 3:
          if stockList[ws.cell(i, 17).value] > int(ws.cell(row=i, column=8).value):
            impendingPrdList.append("○ {} / 상품번호 : {} / 판매상태 : {} / 재고수량 : {} / 데이터파일 기준 재고 : {}".format(ws.cell(i, 17).value, ws.cell(i, 2).value, ws.cell(i, 6).value, ws.cell(i, 8).value, stockList[ws.cell(i, 17).value]))
            
    if ws.cell(row=i, column=6).value == "판매진행":
      if int(ws.cell(row=i, column=8).value) != 0:
        
        # 판매세팅 상품정보 추가
        if ws.cell(i, 14).value not in settingInfo:
          settingInfo[ws.cell(i, 14).value] = [channelName]
        else:
          if channelName not in settingInfo[ws.cell(i, 14).value]:
            settingInfo[ws.cell(i, 14).value].append(channelName)
            
        if prdDetailInfoProduct in excProducts:
          excProductsCheckList.append("○ {} / 상품번호 : {} / 판매상태 : {} / 재고수량 : {}".format(ws.cell(i, 17).value, ws.cell(i, 2).value, ws.cell(i, 6).value, ws.cell(i, 8).value))
  except Exception as e:
    matchingErrList.append('row : {} / {}'.format(i, e))
    continue
  
if len(stockErrList) > 0 or len(stockErrAutoList) > 0:
  f = open("(키디키디) 품절상품 중 판매세팅된 상품 정보.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(키디키디) 품절상품 중 판매세팅된 상품 정보\n\n")
  for i in stockErrList:
    f.write("{}\n\n".format(i))
  for i in stockErrAutoList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(impendingPrdList) > 0:
  f = open("(키디키디) 재고 보충 필요 상품 정보(품절 혹은 품절임박).txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(키디키디) 재고 보충 필요 상품 정보(품절 혹은 품절임박)\n\n")
  for i in impendingPrdList:
    f.write("{}\n\n".format(i))
  f.close()

if len(excProductsCheckList) > 0:
  f = open("(키디키디) 판매제외 상품 포함 체크.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(키디키디) 판매제외 상품 포함 체크\n\n")
  for i in excProductsCheckList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(matchingErrList) > 0:
  f = open("(키디키디) 상품정보 매칭 오류건.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(키디키디) 상품정보 매칭 오류건\n\n")
  for i in matchingErrList:
    f.write("{}\n\n".format(i))
  f.close()

# 판매세팅 상품정보 JSON 파일 업데이트
with open(settingInfoJSONPath, 'w', encoding='UTF-8') as jsonFile:
  json.dump(settingInfo, jsonFile, indent=2, ensure_ascii=False)

wb.active.auto_filter.ref = "A1:R1"
wb.save('상품옵션별 재고현황 추출.xlsx')