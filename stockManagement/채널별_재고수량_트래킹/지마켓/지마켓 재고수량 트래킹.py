from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import productsData
# import os
# from os import listdir
# from os.path import exists
# from os import makedirs

# 재고정보 생성
wbStock = load_workbook('데이터.xlsx')

stockList = {} # 재고정보
stockErrList = [] # 품절상품 중 판매세팅된 상품정보

soldoutPrdCSList = [] # 품절상품(CS팀전달)

impendingPrdList = [] # 재고 보충 필요 상품정보

first_row_cs = 3
last_row_cs = wbStock['품절상품(CS팀전달)'].max_row + 1

for i in range(first_row_cs, last_row_cs):
  soldoutPrdCSList.append(wbStock['품절상품(CS팀전달)'].cell(i, 17).value)

for wbSheet in wbStock:
  wbFirstCell = 3
  wbLastCell = wbSheet.max_row + 1
  
  for i in range(wbFirstCell, wbLastCell):
    if wbSheet.cell(i, 13).value != None:
      stockList[wbSheet.cell(i, 13).value] = wbSheet.cell(i, 14).value

# 상품정보 리스트
product_list = productsData.product_list
color_list = productsData.color_list
size_list = productsData.size_list

wb = load_workbook('./옵션.xlsx')
ws = wb.active

fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillData2 = PatternFill(fill_type='solid', start_color='FFBDBD', end_color='FFBDBD')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True, size=9)

ws.cell(6, 17).value = '상품정보'
ws.cell(6, 18).value = '상품명'
ws.cell(6, 19).value = '컬러'
ws.cell(6, 20).value = '사이즈'
ws.cell(6, 21).value = '주문정보 정제'
ws.cell(6, 22).value = '재고(데이터파일 기준)'

ws.cell(6, 17).alignment = fillAlignment
ws.cell(6, 18).alignment = fillAlignment
ws.cell(6, 19).alignment = fillAlignment
ws.cell(6, 20).alignment = fillAlignment
ws.cell(6, 21).alignment = fillAlignment
ws.cell(6, 22).alignment = fillAlignment

ws.cell(6, 17).font = fillFont
ws.cell(6, 18).font = fillFont
ws.cell(6, 19).font = fillFont
ws.cell(6, 20).font = fillFont
ws.cell(6, 21).font = fillFont
ws.cell(6, 22).font = fillFont

ws.cell(6, 17).fill = fillData
ws.cell(6, 18).fill = fillData
ws.cell(6, 19).fill = fillData
ws.cell(6, 20).fill = fillData
ws.cell(6, 21).fill = fillData
ws.cell(6, 22).fill = fillData

ws.column_dimensions['Q'].width = 40
ws.column_dimensions['R'].width = 25
ws.column_dimensions['S'].width = 25
ws.column_dimensions['T'].width = 25
ws.column_dimensions['U'].width = 40
ws.column_dimensions['V'].width = 25


first_row = 7
last_row = ws.max_row + 1

for i in range(first_row, last_row):
  try:
    ws.cell(i, 17).value = str(ws.cell(i, 6).value) + '/' + str(ws.cell(i, 7).value)
    
    for product in product_list:
      try:
        if product in str(ws.cell(row=i, column=17).value):
          prdDetailInfoProduct = product.replace("(저스틴23)", "")
          ws.cell(row=i, column=18).value = prdDetailInfoProduct
      except:
        pass
    for color in color_list:
      if color in str(ws.cell(row=i, column=17).value):
        prdDetailInfoColor = color
        ws.cell(row=i, column=19).value = prdDetailInfoColor
    for size in size_list:
      if size in str(ws.cell(row=i, column=17).value):
        prdDetailInfoSize = size.replace("FREE", "free")
        ws.cell(row=i, column=20).value = prdDetailInfoSize
    
    prdDetailInfo = '{} {} {}'.format(prdDetailInfoProduct, prdDetailInfoColor, prdDetailInfoSize)
    
    ws.cell(i, 21).value = prdDetailInfo
    ws.cell(i, 22).value = stockList[ws.cell(i, 21).value]
    
    if stockList[ws.cell(i, 21).value] == 0:
      if ws.cell(row=i, column=12).value == "Y":
        if int(ws.cell(row=i, column=16).value) != 0 or ws.cell(row=i, column=11).value == "정상":
          print("{}/{}".format(ws.cell(i, 21).value, stockList[ws.cell(i, 21).value]))
          if ws.cell(i, 21).value in soldoutPrdCSList:
            stockErrList.append("○ {} / 상태 : {} / 노출여부 : {} / 재고수량 : {} / 데이터파일 기준 재고 : 0".format(ws.cell(i, 21).value, ws.cell(i, 11).value, ws.cell(i, 12).value, ws.cell(i, 16).value))
          else:
            stockErrList.append("※ 판매량차감 자동품절 상품(CS팀에서 품절로 전달되지 않은 상품) ※\n○ {} / 상태 : {} / 노출여부 : {} / 재고수량 : {} / 데이터파일 기준 재고 : 0".format(ws.cell(i, 21).value, ws.cell(i, 11).value, ws.cell(i, 12).value, ws.cell(i, 16).value))
          for colNum in range(1, 23):
            ws.cell(row=i, column=colNum).fill = fillData2
            
    if stockList[ws.cell(i, 21).value] != 0:
      if int(ws.cell(row=i, column=16).value) <= 3:
        impendingPrdList.append("○ {} / 상태 : {} / 노출여부 : {} / 재고수량 : {} / 데이터파일 기준 재고 : {}".format(ws.cell(i, 21).value, ws.cell(i, 11).value, ws.cell(i, 12).value, ws.cell(i, 16).value, stockList[ws.cell(i, 21).value]))
        
  except:
    continue
  
if len(stockErrList) > 0:
  f = open("(지마켓) 품절상품 중 판매세팅된 상품 정보.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(지마켓) 품절상품 중 판매세팅된 상품 정보\n\n")
  for i in stockErrList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(impendingPrdList) > 0:
  f = open("(지마켓) 재고 보충 필요 상품 정보(품절 혹은 품절임박).txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(지마켓) 재고 보충 필요 상품 정보(품절 혹은 품절임박)\n\n")
  for i in impendingPrdList:
    f.write("{}\n\n".format(i))
  f.close()  

wb.active.auto_filter.ref = "A6:V6"
wb.save('상품옵션별 재고현황 추출.xlsx')