from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import os
from datetime import datetime

wb = load_workbook('판매데이터.xlsx')
sheet1Name = '판매량 체크'

currPath = os.getcwd()
now = datetime.now()

fillData = PatternFill(fill_type='solid', start_color='FF7789', end_color='FF7789')
fillData2 = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')

firstCell = 2
lastCell = wb[sheet1Name].max_row + 1

# 모션 옵션리스트
prdCapOptionList = [
  '토밍이세트/아이보리/free',
  '토밍이세트/브라운/free',
  '토밍이세트/블랙/free',
  '토밍이세트/핑크/free',
  '보스턴비니/블랙/free',
  '보스턴비니/아이보리/free',
  '보스턴비니/그레이그린/free',
  '코코모자/카키/free',
  '코코모자/그레이/free',
  '코코모자/브라운/free',
  '토끼요정모자/베이지/free',
  '토끼요정모자/핑크/free',
  '토끼요정모자/카키/free',
  '토끼요정모자/그레이/free',
  '티라노모자/아이보리/free',
  '티라노모자/레드/free',
  '티라노모자/블랙/free',
  '티라노모자/오렌지/free',
  '해피스노우세트/브라운/free',
  '해피스노우세트/아이보리/free',
  '해피스노우세트/핑크/free',
  '퍼피캡모자/카라멜/free',
  '퍼피캡모자/블루/free',
  '퍼피캡모자/오렌지/free',
  '퍼피캡모자/그린/free',
  '퍼피캡모자/네이비/free',
  '퍼피모자/아이보리/free',
  '퍼피모자/핑크/free',
  '퍼피모자/오렌지/free',
  '퍼피모자/옐로우/free',
  '퍼피모자/민트/free',
  '퍼피모자/블루/free',
  '카우보이별모자/진청/free',
  '곰돌이모자/아이보리/free',
  '곰돌이모자/베이지/free',
  '곰돌이모자/블랙/free',
  '곰돌이모자/브라운/free',
  '용용이모자/베이지/free',
  '용용이모자/핑크/free',
  '용용이모자/옐로우/free',
  '용용이모자/네이비/free',
  '베어캡모자/그린/free',
  '베어캡모자/옐로우/free',
  '베어캡모자/아이보리/free',
  '베어캡모자/오렌지/free',
  '베어캡모자/네이비/free',
  '알파벳모자/핑크/free',
  '알파벳모자/옐로우/free',
  '알파벳모자/화이트/free',
  '알파벳모자/블루/free',
  '알파벳모자/오렌지/free',
  '알파벳벙거지/화이트/free',
  '알파벳벙거지/핑크/free',
  '알파벳벙거지/아이보리/free',
  '알파벳벙거지/블랙/free',
  '알파벳벙거지/블루/free',
  '보스턴챙모자/베이지/free',
  '보스턴챙모자/아이보리/free',
  '보스턴챙모자/브라운/free',
  '보스턴챙모자/블랙/free',
  '스마일비니/옐로우/free',
  '스마일비니/그레이/free',
  '스마일비니/블랙/free',
  '스마일비니/베이지/free',
  '스마일비니/레드/free',
  '스마일비니/핑크/free',
  '스마일비니/블루/free',
  '스마일비니/브라운/free',
]

# 모자판매정보 생성
capSellList = {}
for i in range(firstCell, lastCell):
  try:
    if wb[sheet1Name].cell(i, 1).value in prdCapOptionList:
      capSellList[wb[sheet1Name].cell(i, 1).value] = wb[sheet1Name].cell(i, 2).value
  except:
    pass
  
if len(capSellList) > 0:
  f = open('모자상품_판매량추출.txt', 'w')
  for key, value in capSellList.items():
    f.write('{}/{}\n'.format(key, value))
  f.close()