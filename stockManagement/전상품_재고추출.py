from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import os

wb = load_workbook('데이터.xlsx')

currPath = os.getcwd()

# 재고정보 생성
stockList = {}

for wbSheet in wb:
  wbFirstCell = 3
  wbLastCell = wbSheet.max_row + 1
  
  for i in range(wbFirstCell, wbLastCell):
    if wbSheet.cell(i, 13).value != None:
      stockList[wbSheet.cell(i, 13).value] = wbSheet.cell(i, 14).value

wbStock = Workbook()
wbStockSheet = wbStock.active

fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True)

wbStockSheet.cell(1, 1).alignment = fillAlignment
wbStockSheet.cell(1, 2).alignment = fillAlignment
wbStockSheet.cell(1, 1).font = fillFont
wbStockSheet.cell(1, 2).font = fillFont
wbStockSheet.cell(1, 1).fill = fillData
wbStockSheet.cell(1, 2).fill = fillData

wbStockSheet.column_dimensions['A'].width = 60
wbStockSheet.column_dimensions['B'].width = 25
wbStockSheet.cell(1, 1).value = "상품식별값"
wbStockSheet.cell(1, 2).value = "재고수량"

print(stockList)

for key, value in stockList.items():
  wbStockRow = wbStockSheet.max_row + 1
  wbStockSheet.cell(wbStockRow, 1).value = key
  wbStockSheet.cell(wbStockRow, 2).value = value

wbStock.save(currPath + '\\전상품_재고추출.xlsx')


# firstCell = 2
# lastCell = wb[sheet1Name].max_row + 1


# for i in range(firstCell, lastCell):
#   sellList[wb[sheet1Name].cell(i, 1).value.replace('(저스틴23)', '').replace('/', ' ')] = wb[sheet1Name].cell(i, 2).value

# print(sellList)

# # 전체 재고리스트
# prdList = []

# # 판매량 차감 후 품절된 상품리스트
# prdSoldoutList = []

# # 중복상품 체크용 딕셔너리
# doublePrdList = {}

# wb2 = load_workbook('데이터.xlsx')

# for wb2Sheet in wb2:
#   wb2FirstCell = 3
#   wb2LastCell = wb2Sheet.max_row + 1
  
#   for i in range(wb2FirstCell, wb2LastCell):
#     if wb2Sheet.cell(i, 13).value != None:
#       prdList.append(wb2Sheet.cell(i, 13).value)
      
#     # 중복상품 체크
#     if wb2Sheet.cell(i, 13).value not in doublePrdList and wb2Sheet.cell(i, 13).value != None:
#       doublePrdList[wb2Sheet.cell(i, 13).value] = 1
#     elif wb2Sheet.cell(i, 13).value in doublePrdList and wb2Sheet.cell(i, 13).value != None:
#       doublePrdList[wb2Sheet.cell(i, 13).value] += 1
      
#     if sellList.get(wb2Sheet.cell(i, 13).value) != None:
#       if wb2Sheet.cell(i, 14).value > 0:
#         if sellList[wb2Sheet.cell(i, 13).value] > wb2Sheet.cell(i, 14).value:
#           wb2Sheet.cell(i, 14).value = 0
#           prdSoldoutList.append(str(wb2Sheet.cell(i, 13).value) + '/' + str(now.strftime('%Y-%m-%d')))
#         else:  
#           wb2Sheet.cell(i, 14).value -= sellList[wb2Sheet.cell(i, 13).value]
#       else:
#         wb2Sheet.cell(i, 14).value = 0
#     else:
#       continue
    
# # 차감 재고 데이터 매칭 오류 정보 저장
# errList = []

# for i in sellList:
#   if i not in prdList:
#     print('{} : {}'.format(i, sellList[i]))
#     errList.append(i)

# if len(errList) > 0:
#   f = open('차감 재고 데이터 매칭 오류.txt', 'w')
#   for i in errList:
#     f.write('{} : {}\n'.format(i, sellList[i]))
#   f.close()

# # 판매량 차감 후 품절처리된 상품 정보 저장
# if len(prdSoldoutList) > 0:
#   f2 = open('판매량 차감 후 품절처리된 상품 정보.txt', 'w')
#   for i in prdSoldoutList:
#     f2.write('{}\n'.format(i))
#   f2.close()

# # 중복상품 체크 정보 저장
# isDouble = False
# for i in doublePrdList.values():
#   if i > 1:
#     isDouble = True

# if isDouble:
#   f3 = open('중복상품체크.txt', 'w')
#   for key, value in doublePrdList.items():
#     if key is not None and value > 1:
#       f3.write('{} : {}\n'.format(key, value))
#   f3.close()

# wb2.save(currPath + '\\데이터_판매량 반영.xlsx')