from openpyxl import load_workbook

wb = load_workbook('판매데이터.xlsx')
sheet1Name = '판매량 체크'

firstCell = 2
lastCell = wb[sheet1Name].max_row + 1

# 모자 옵션리스트
prdCapOptionList = [
  '스마일비니/옐로우/free',
  '스마일비니/그레이/free',
  '스마일비니/블랙/free',
  '스마일비니/베이지/free',
  '스마일비니/레드/free',
  '스마일비니/핑크/free',
  '스마일비니/블루/free',
  '스마일비니/브라운/free',
]

# 모자 판매채널 정보
prdCapChannelList = {}

for i in range(firstCell, lastCell):
  prdCapChannelList[wb[sheet1Name].cell(i, 25).value] = wb[sheet1Name].cell(i, 26).value

# 모자판매정보 생성
capSellList = {}
for i in range(firstCell, lastCell):
  try:
    if wb[sheet1Name].cell(i, 1).value in prdCapOptionList:
      capSellList[wb[sheet1Name].cell(i, 1).value] = str(prdCapChannelList[wb[sheet1Name].cell(i, 1).value]) + " / " + str(wb[sheet1Name].cell(i, 2).value)
  except:
    pass
  
if len(capSellList) > 0:
  f = open('스마일비니_판매량추출.txt', 'w')
  for key, value in capSellList.items():
    f.write('{}/{}\n'.format(key, value))
  f.close()