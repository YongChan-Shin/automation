from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import os
from datetime import datetime

wb = load_workbook('판매데이터.xlsx')
sheet1Name = '판매량 체크'

currPath = os.getcwd()
now = datetime.now()

fillData = PatternFill(fill_type='solid', start_color='FF7789', end_color='FF7789')
fillData2 = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')

firstCell = 2
lastCell = wb[sheet1Name].max_row + 1

# 판매정보 생성
sellList = {}
for i in range(firstCell, lastCell):
  try:
    sellList[wb[sheet1Name].cell(i, 1).value.replace('(저스틴23)', '').replace('/', ' ')] = wb[sheet1Name].cell(i, 2).value
  except:
    pass
  
# 판매채널정보 생성
sellChannelList ={}
for i in range(firstCell, lastCell):
  try:
    sellChannelList[wb[sheet1Name].cell(i, 22).value] = wb[sheet1Name].cell(i, 23).value
  except:
    pass
  
# 판매채널정보(상세) 생성
sellChannelDetailList ={}
for i in range(firstCell, lastCell):
  try:
    sellChannelDetailList[wb[sheet1Name].cell(i, 25).value.replace("/", " ")] = wb[sheet1Name].cell(i, 26).value
  except:
    pass
  
# 상품별 주문번호정보 생성
# orderNumList ={}
# for i in range(firstCell, lastCell):
#   try:
#     orderNumList[wb[sheet1Name].cell(i, 25).value] = wb[sheet1Name].cell(i, 26).value
#   except:
#     pass
  
# 상품별(컬러/사이즈) 주문번호정보 생성
orderNumDetailList ={}
for i in range(firstCell, lastCell):
  try:
    orderNumDetailList[wb[sheet1Name].cell(i, 31).value.replace("/", " ")] = wb[sheet1Name].cell(i, 32).value
  except:
    pass

print(sellList)

# 전체 재고상품명 리스트
prdList = []

# 재고정보 생성
stockList = {}

# 재고정보(사이즈종합) 생성
stockAllSizeList = {}
currentPrd = ""

# 재고수량(사이즈종합) 30개 미만 상품 리스트
stockImpendingCheckSheetName = "겨울_이월" # TODO 확인 후 수정필요
stockImpendingCheckSheetName2 = "" # TODO 확인 후 수정필요
stockImpending = []

# 품절상품 중 재고수량 0인 아닌 상품 식별 리스트
stockErrList = []

# 품절상품(CS팀전달)
soldoutPrdCSList = []

# 재고수량 0인 상품 중 판매된 상품리스트
prdOrderSoldout = []
# 재고수량 0인 상품 중 판매된 상품리스트(자동품절건 - CS팀에서 품절 넘어오지 않은 상품)
prdOrderSoldoutAuto = []

# 판매량 차감 후 품절된 상품리스트
prdSoldoutList = []

# 중복상품 체크용 딕셔너리
doublePrdList = {}

# 세팅채널과 판매채널 상이한 상품리스트
channelErrPrdList = []

# 별이오키즈 상품리스트
byeorioPrdList = []

# DB 불러오기
import sqlite3
con = sqlite3.connect('D:/1.업무/10.기타자료/Development/db/productsData.db')
cur = con.cursor()

cur.execute("SELECT PrdName from ProductsData WHERE SeasonDetail like 'BYEORIO'")
data = cur.fetchall()
for i in data:
  byeorioPrdList.append(i[0])
  
wb2 = load_workbook('데이터.xlsx')

first_row_cs = 3
last_row_cs = wb2['품절상품(CS팀전달)'].max_row + 1

for i in range(first_row_cs, last_row_cs):
  soldoutPrdCSList.append(wb2['품절상품(CS팀전달)'].cell(i, 17).value)

for wb2Sheet in wb2:
  wb2FirstCell = 3
  wb2LastCell = wb2Sheet.max_row + 1
  
  for i in range(wb2FirstCell, wb2LastCell):
    try:
      if wb2Sheet.cell(i, 13).value != None:
        prdList.append(wb2Sheet.cell(i, 13).value)
        stockList[wb2Sheet.cell(i, 13).value] = wb2Sheet.cell(i, 14).value
    except:
      pass
      
    # 중복상품 체크
    if wb2Sheet.cell(i, 13).value not in doublePrdList and wb2Sheet.cell(i, 13).value != None:
      doublePrdList[wb2Sheet.cell(i, 13).value] = 1
    elif wb2Sheet.cell(i, 13).value in doublePrdList and wb2Sheet.cell(i, 13).value != None:
      doublePrdList[wb2Sheet.cell(i, 13).value] += 1
      
    # 판매상품 재고 차감
    try:
      if wb2Sheet.cell(i, 13).value.split(' ')[0] not in byeorioPrdList:
        continue
      else:
        if sellList.get(wb2Sheet.cell(i, 13).value) != None:
          if wb2Sheet.cell(i, 14).value > 0:
            if sellList[wb2Sheet.cell(i, 13).value] >= wb2Sheet.cell(i, 14).value:
              prdSoldoutList.append(str(wb2Sheet.cell(i, 13).value) + '/' + str(sellList[wb2Sheet.cell(i, 13).value]) + '개판매/' + '차감 후 수량 : ' + str(wb2Sheet.cell(i, 14).value - sellList[wb2Sheet.cell(i, 13).value]) + '개/' + str(now.strftime('%Y-%m-%d')) + '\n- 세팅채널 : ' + str(wb2Sheet.cell(i, 19).value) + '\n- 판매채널 : ' + str(sellChannelDetailList[wb2Sheet.cell(i, 13).value]) + '\n- 주문번호 : ' + str(orderNumDetailList[wb2Sheet.cell(i, 13).value]))
              wb2Sheet.cell(i, 14).value = 0
            else:
              wb2Sheet.cell(i, 14).value -= sellList[wb2Sheet.cell(i, 13).value]
          else:
            if wb2Sheet.cell(i, 13).value not in soldoutPrdCSList:
              prdOrderSoldoutAuto.append("(※ 판매량차감 자동품절 상품(CS팀에서 품절로 전달되지 않은 상품) ※) " + str(wb2Sheet.cell(i, 13).value) + '/' + str(sellList[wb2Sheet.cell(i, 13).value]) + '개판매' + '\n- 세팅채널 : ' + str(wb2Sheet.cell(i, 19).value) + '\n- 판매채널 : ' + str(sellChannelDetailList[wb2Sheet.cell(i, 13).value]) + '\n- 주문번호 : ' + str(orderNumDetailList[wb2Sheet.cell(i, 13).value]))
            else:  
              prdOrderSoldout.append(str(wb2Sheet.cell(i, 13).value) + '/' + str(sellList[wb2Sheet.cell(i, 13).value]) + '개판매' + '\n- 세팅채널 : ' + str(wb2Sheet.cell(i, 19).value) + '\n- 판매채널 : ' + str(sellChannelDetailList[wb2Sheet.cell(i, 13).value]) + '\n- 주문번호 : ' + str(orderNumDetailList[wb2Sheet.cell(i, 13).value]))
            wb2Sheet.cell(i, 14).value = 0
            
          # 누적판매량 정리
          if wb2Sheet.cell(i, 16).value != None:
            wb2Sheet.cell(i, 16).value += sellList[wb2Sheet.cell(i, 13).value]
          else:
            wb2Sheet.cell(i, 16).value = sellList[wb2Sheet.cell(i, 13).value]
        else:
          pass
    except:
      pass

    # 사이즈 종합 재고수량 정보 저장
    try:
      if wb2Sheet.cell(i, 13).value != None:
        if wb2Sheet.title == stockImpendingCheckSheetName or wb2Sheet.title == stockImpendingCheckSheetName2:
          if wb2Sheet.cell(i, 5).value != currentPrd:
            currentPrd = wb2Sheet.cell(i, 5).value
            stockAllSizeList[currentPrd] = wb2Sheet.cell(i, 14).value
          else:
            stockAllSizeList[currentPrd] += wb2Sheet.cell(i, 14).value
    except:
      pass
    
    # # 실제 판매채널 체크
    # if wb2Sheet.cell(i, 5).value != None:
    #   wb2Sheet.cell(i, 20).value = " "
    #   try:
    #     wb2Sheet.cell(i, 20).value = sellChannelList[wb2Sheet.cell(i, 5).value]
    #     if wb2Sheet.cell(i, 19).value != None:
    #       for prd in wb2Sheet.cell(i, 20).value.split('/'):
    #         # 세팅채널과 판매채널 상이 시 특이사항 표시
    #         if sellList.get(wb2Sheet.cell(i, 13).value) > 0 and prd not in wb2Sheet.cell(i, 19).value:
    #           wb2Sheet.cell(i, 19).fill = fillData
    #           wb2Sheet.cell(i, 19).value = '(' + prd + ')' + wb2Sheet.cell(i, 19).value
    #           wb2Sheet.cell(i, 20).fill = fillData
    #           if str(wb2Sheet.cell(i, 4).value+"/"+wb2Sheet.cell(i, 5).value) not in channelErrPrdList:
    #             channelErrPrdList.append(str(wb2Sheet.cell(i, 4).value+"/"+wb2Sheet.cell(i, 5).value))
    #   except:
    #     pass
      
    # 실제 판매채널(상세) 체크
    if wb2Sheet.cell(i, 13).value != None:
      wb2Sheet.cell(i, 20).value = " "
      try:
        wb2Sheet.cell(i, 20).value = sellChannelDetailList[wb2Sheet.cell(i, 13).value]
        if wb2Sheet.cell(i, 19).value != None:
          for channel in wb2Sheet.cell(i, 20).value.split('/'):
            # 세팅채널과 판매채널 상이 시 특이사항 표시
            if sellList.get(wb2Sheet.cell(i, 13).value) > 0 and channel not in wb2Sheet.cell(i, 19).value:
              wb2Sheet.cell(i, 19).fill = fillData
              wb2Sheet.cell(i, 19).value = '(' + channel + ')' + wb2Sheet.cell(i, 19).value
              wb2Sheet.cell(i, 20).fill = fillData
              if str(wb2Sheet.cell(i, 4).value+"/"+wb2Sheet.cell(i, 5).value) not in channelErrPrdList:
                channelErrPrdList.append(str(wb2Sheet.cell(i, 4).value+"/"+wb2Sheet.cell(i, 5).value))
        wb2Sheet.cell(i, 20).value = "({}개 판매) {} : {}".format(sellList[wb2Sheet.cell(i, 13).value], sellChannelDetailList[wb2Sheet.cell(i, 13).value], orderNumDetailList[wb2Sheet.cell(i, 13).value])
      except:
        pass
    
# 차감 재고 데이터 매칭 오류 정보 저장
errList = []

for i in sellList:
  if i.split(' ')[0] in byeorioPrdList and  i not in prdList:
    print('{} : {} / {}'.format(i, sellList[i], sellChannelList[i.split(' ')[0]]))
    errList.append(i)

if len(errList) > 0 or len(prdOrderSoldout) > 0 or len(prdOrderSoldoutAuto) > 0:
  f = open('차감 재고 데이터 매칭 오류.txt', 'w')
  for i in errList:
    if i.split(' ')[0] in byeorioPrdList:
      continue
    else:
      f.write('{} : {} / {}\n'.format(i, sellList[i], sellChannelList[i.split(" ")[0]]))
  f.write('\n\n\nㅡㅡㅡㅡㅡㅡㅡㅡㅡ 재고수량 0인 상품 중 판매된 상품리스트 ㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n')
  f.close()
  f = open('차감 재고 데이터 매칭 오류.txt', 'a')
  for i in prdOrderSoldout:
    f.write('○ {} \n\n'.format(i))
  for i in prdOrderSoldoutAuto:
    f.write('○ {} \n\n'.format(i))
  f.close()
  
# 판매량 차감 후 품절처리된 상품 정보 저장
if len(prdSoldoutList) > 0:
  f2 = open('판매량 차감 후 품절처리된 상품 정보.txt', 'w')
  f2.write('ㅡㅡㅡㅡㅡㅡㅡㅡㅡ 판매량 차감 후 품절처리된 상품 정보 ㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n')
  for i in prdSoldoutList:
    f2.write('○ {}\n\n'.format(i))
  f2.close()

# 중복상품 체크 정보 저장
isDouble = False
for i in doublePrdList.values():
  if i > 1:
    isDouble = True

if isDouble:
  f3 = open('중복상품체크.txt', 'w')
  for key, value in doublePrdList.items():
    if key is not None and value > 1:
      f3.write('{} : {}\n'.format(key, value))
  f3.close()
  
wb2Soldout1FirstCell = 3
wb2Soldout1LastCell = wb2['품절상품(CS팀전달)'].max_row + 1

wb2Soldout2FirstCell = 3
wb2Soldout2LastCell = wb2['품절상품(판매량차감)'].max_row + 1

for i in range(wb2Soldout1FirstCell, wb2Soldout1LastCell):
  try:
    if stockList[wb2['품절상품(CS팀전달)'].cell(i, 17).value] != 0:
      stockErrList.append('{} / {}'.format(wb2['품절상품(CS팀전달)'].cell(i, 17).value, stockList[wb2['품절상품(CS팀전달)'].cell(i, 17).value]))
  except:
    pass

for i in range(wb2Soldout2FirstCell, wb2Soldout2LastCell):
  try:
    if stockList[wb2['품절상품(판매량차감)'].cell(i, 17).value] != 0:
      stockErrList.append('{} / {}'.format(wb2['품절상품(판매량차감)'].cell(i, 17).value, stockList[wb2['품절상품(판매량차감)'].cell(i, 17).value]))
  except:
    pass
  
# 사이즈 종합 재고수량 30개 미만 상품 정보 저장
for key, value in stockAllSizeList.items():
  try:
    if value >= 0 and value < 30:
      stockImpending.append("\n○ {}/총 재고 : {}".format(key, value))
      for wb2Sheet in wb2:
        if wb2Sheet.title == stockImpendingCheckSheetName or wb2Sheet.title == stockImpendingCheckSheetName2:
          wb2FirstCell = 3
          wb2LastCell = wb2Sheet.max_row + 1
          for i in range(wb2FirstCell, wb2LastCell):
            if wb2Sheet.cell(i, 5).value == key:
              stockImpending.append("- {} : {}".format(wb2Sheet.cell(i, 13).value, wb2Sheet.cell(i, 14).value))
  except:
    pass

if len(stockErrList) > 0:
  f4 = open('품절상품 중 재고수량 0인 아닌 상품 정보.txt', 'w')
  for i in stockErrList:
    f4.write('{}\n'.format(i))
  f4.close()

if len(channelErrPrdList) > 0:
  f5 = open('세팅채널과 판매채널이 상이한 상품 정보.txt', 'w')
  for i in channelErrPrdList:
    f5.write('{}\n'.format(i))
  f5.close()
  
if len(stockImpending) > 0:
  f6 = open('전 사이즈 총 재고수량 30개 미만 상품.txt', 'w')
  f6.write('ㅡㅡㅡㅡㅡㅡㅡㅡ 시트 : {} / {} ㅡㅡㅡㅡㅡㅡㅡㅡ\n'.format(stockImpendingCheckSheetName, stockImpendingCheckSheetName2)) # TODO 확인 후 수정필요
  for i in stockImpending:
    f6.write('{}\n'.format(i))
  f6.close()

wb2.save(currPath + '\\데이터_판매량 반영.xlsx')