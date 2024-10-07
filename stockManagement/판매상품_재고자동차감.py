from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import os
from datetime import datetime

wb = load_workbook('판매데이터.xlsx')
sheet1Name = '판매량 체크'

currPath = os.getcwd()
now = datetime.now()

fillData = PatternFill(fill_type='solid', start_color='FF7789', end_color='FF7789')
fillData2 = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')

firstCell = 2
lastCell = wb[sheet1Name].max_row + 1

# 판매정보 생성
sellList = {}
for i in range(firstCell, lastCell):
  try:
    sellList[wb[sheet1Name].cell(i, 1).value.replace('(저스틴23)', '').replace('/', ' ')] = wb[sheet1Name].cell(i, 2).value
  except:
    pass
  
# 판매채널정보 생성
sellChannelList ={}
for i in range(firstCell, lastCell):
  try:
    sellChannelList[wb[sheet1Name].cell(i, 22).value] = wb[sheet1Name].cell(i, 23).value
  except:
    pass

print(sellList)

# 전체 재고상품명 리스트
prdList = []

# 재고정보 생성
stockList = {}

# 품절상품 중 재고수량 0인 아닌 상품 식별 리스트
stockErrList = []

# 재고수량 0인 상품 중 판매된 상품리스트
prdOrderSoldout = []

# 판매량 차감 후 품절된 상품리스트
prdSoldoutList = []

# 중복상품 체크용 딕셔너리
doublePrdList = {}

wb2 = load_workbook('데이터.xlsx')

for wb2Sheet in wb2:
  wb2FirstCell = 3
  wb2LastCell = wb2Sheet.max_row + 1
  
  for i in range(wb2FirstCell, wb2LastCell):
    if wb2Sheet.cell(i, 13).value != None:
      prdList.append(wb2Sheet.cell(i, 13).value)
      stockList[wb2Sheet.cell(i, 13).value] = wb2Sheet.cell(i, 14).value
      
    # 중복상품 체크
    if wb2Sheet.cell(i, 13).value not in doublePrdList and wb2Sheet.cell(i, 13).value != None:
      doublePrdList[wb2Sheet.cell(i, 13).value] = 1
    elif wb2Sheet.cell(i, 13).value in doublePrdList and wb2Sheet.cell(i, 13).value != None:
      doublePrdList[wb2Sheet.cell(i, 13).value] += 1
      
    # 판매상품 재고 차감
    if sellList.get(wb2Sheet.cell(i, 13).value) != None:
      if wb2Sheet.cell(i, 14).value > 0:
        if sellList[wb2Sheet.cell(i, 13).value] > wb2Sheet.cell(i, 14).value:
          prdSoldoutList.append(str(wb2Sheet.cell(i, 13).value) + '/' + str(wb2Sheet.cell(i, 14).value - sellList[wb2Sheet.cell(i, 13).value]) + '/' + str(now.strftime('%Y-%m-%d')))
          wb2Sheet.cell(i, 14).value = 0
        else:
          wb2Sheet.cell(i, 14).value -= sellList[wb2Sheet.cell(i, 13).value]
      else:
        prdOrderSoldout.append((str(wb2Sheet.cell(i, 13).value) + '/' + str(sellList[wb2Sheet.cell(i, 13).value])))
        wb2Sheet.cell(i, 14).value = 0
        
      # 누적판매량 정리
      if wb2Sheet.cell(i, 16).value != None:
        wb2Sheet.cell(i, 16).value += sellList[wb2Sheet.cell(i, 13).value]
      else:
        wb2Sheet.cell(i, 16).value = sellList[wb2Sheet.cell(i, 13).value]
    else:
      pass
    
    # 실제 판매채널 체크
    if wb2Sheet.cell(i, 5).value != None:
      wb2Sheet.cell(i, 20).value = " "
      try:
        wb2Sheet.cell(i, 20).value = sellChannelList[wb2Sheet.cell(i, 5).value]
        if wb2Sheet.cell(i, 19).value != None:
          for prd in wb2Sheet.cell(i, 20).value.split('/'):
            if prd not in wb2Sheet.cell(i, 19).value:
              wb2Sheet.cell(i, 19).fill = fillData
              wb2Sheet.cell(i, 19).value = '(' + prd + ')' + wb2Sheet.cell(i, 19).value
              wb2Sheet.cell(i, 20).fill = fillData
      except:
        pass
    
# 차감 재고 데이터 매칭 오류 정보 저장
errList = []

for i in sellList:
  if i not in prdList:
    print('{} : {} / {}'.format(i, sellList[i], sellChannelList[i.split(" ")[0]]))
    errList.append(i)

if len(errList) > 0 or len(prdOrderSoldout) > 0:
  f = open('차감 재고 데이터 매칭 오류.txt', 'w')
  for i in errList:
    f.write('{} : {} / {}\n'.format(i, sellList[i], sellChannelList[i.split(" ")[0]]))
  f.write('\n\n\nㅡㅡㅡㅡㅡㅡㅡㅡㅡ 재고수량 0인 상품 중 판매된 상품리스트 ㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n')
  f.close()
  f = open('차감 재고 데이터 매칭 오류.txt', 'a')
  for i in prdOrderSoldout:
    f.write('{} / {}\n'.format(i, sellChannelList[i.split(" ")[0]]))
  f.close()

# 판매량 차감 후 품절처리된 상품 정보 저장
if len(prdSoldoutList) > 0:
  f2 = open('판매량 차감 후 품절처리된 상품 정보.txt', 'w')
  for i in prdSoldoutList:
    f2.write('{}\n'.format(i))
  f2.close()

# 중복상품 체크 정보 저장
isDouble = False
for i in doublePrdList.values():
  if i > 1:
    isDouble = True

if isDouble:
  f3 = open('중복상품체크.txt', 'w')
  for key, value in doublePrdList.items():
    if key is not None and value > 1:
      f3.write('{} : {}\n'.format(key, value))
  f3.close()
  
wb2Soldout1FirstCell = 3
wb2Soldout1LastCell = wb2['품절상품(CS팀전달)'].max_row + 1

wb2Soldout2FirstCell = 3
wb2Soldout2LastCell = wb2['품절상품(판매량차감)'].max_row + 1

for i in range(wb2Soldout1FirstCell, wb2Soldout1LastCell):
  try:
    if stockList[wb2['품절상품(CS팀전달)'].cell(i, 17).value] != 0:
      stockErrList.append('{} / {}'.format(wb2['품절상품(CS팀전달)'].cell(i, 17).value, stockList[wb2['품절상품(CS팀전달)'].cell(i, 17).value]))
  except:
    pass

for i in range(wb2Soldout2FirstCell, wb2Soldout2LastCell):
  try:
    if stockList[wb2['품절상품(판매량차감)'].cell(i, 17).value] != 0:
      stockErrList.append('{} / {}'.format(wb2['품절상품(판매량차감)'].cell(i, 17).value, stockList[wb2['품절상품(판매량차감)'].cell(i, 17).value]))
  except:
    pass

if len(stockErrList) > 0:
  f4 = open('품절상품 중 재고수량 0인 아닌 상품 정보.txt', 'w')
  for i in stockErrList:
    f4.write('{}\n'.format(i))
  f4.close()

wb2.save(currPath + '\\데이터_판매량 반영.xlsx')