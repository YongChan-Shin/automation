from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

driver.get('https://shop.kidsnote.com/_manage')

url = input('작업 시작(대상 url 입력)')
  
driver.get(url)

addBtn = driver.find_element(By.CSS_SELECTOR, '#manage > div.popupContent > form:nth-child(5) > table > tbody > tr:nth-child(8) > td > div > span > input[type=button]')

wb = load_workbook('data.xlsx')
ws = wb.active
first_row = 2
last_row = ws.max_row + 1

for i in range(first_row, last_row):
  print(ws.cell(i, 1).value, ws.cell(i, 2).value)
  addBtn.click()
  time.sleep(0.4)
  optionRow = driver.find_element(By.ID, 'option_row_{}'.format(i-2))
  optionEl = optionRow.find_elements(By.TAG_NAME, 'td')[0].find_element(By.CLASS_NAME, 'name')
  priceEl = optionRow.find_elements(By.TAG_NAME, 'td')[1].find_element(By.CLASS_NAME, 'add_price')
  optionEl.clear()
  optionEl.send_keys(ws.cell(i, 1).value)
  priceEl.clear()
  priceEl.send_keys(ws.cell(i, 2).value)

time.sleep(100000000)